# trade_log.py
"""AI 策略交易日志 —— 自动将 AI 分析结果写入 Excel"""

import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

# 日志文件路径（项目根目录）
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AI策略交易日志.xlsx")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    _CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL = PatternFill("solid", fgColor="FFC7CE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
    HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BLUE_FONT = Font(name="Arial", color="0066CC", size=10)
    STRONG_FONT = Font(name="Arial", bold=True, size=10)

    def _xlsx_available() -> bool:
        return True

except ImportError:
    _xlsx_available = lambda: False


def is_available() -> bool:
    """检查 openpyxl 是否已安装。"""
    return _xlsx_available()


def _find_next_row(ws) -> int:
    """找到下一个空白行（从第3行开始，B列为信号日）。"""
    for row in range(3, 53):
        if ws.cell(row=row, column=2).value is None:
            return row
    return 53


def write_signal(
    symbol: str,
    direction: str,
    entry: float,
    stop_loss: float,
    take_profit: float,
    margin: float = 10.0,
    leverage: int = 1,
    backtest_expect_pct: float = None,
    note: str = "",
) -> int:
    """
    追加一条 AI 信号到交易日志。

    Args:
        symbol: 交易对，如 "AI-USDT-SWAP"
        direction: "long" / "short" / "neutral"
        entry: AI 入场价
        stop_loss: AI 止损价
        take_profit: AI 止盈价
        margin: 保证金金额 ($)
        leverage: 杠杆倍数
        backtest_expect_pct: 回测预期盈亏 (%)
        note: 备注（AI 的预测简述）

    Returns:
        写入的行号，失败返回 -1
    """
    if not _xlsx_available():
        logger.error("openpyxl 未安装，无法写入交易日志")
        return -1

    if not os.path.exists(LOG_FILE):
        logger.info("日志文件不存在，正在创建: %s", LOG_FILE)
        if not init_log_file():
            return -1

    try:
        wb = load_workbook(LOG_FILE)
    except Exception as e:
        logger.error("打开日志文件失败: %s", e)
        return -1

    if "交易日志" not in wb.sheetnames:
        logger.error("日志文件中没有 '交易日志' Sheet")
        wb.close()
        return -1

    ws = wb["交易日志"]
    row = _find_next_row(ws)

    if row > 52:
        logger.warning("交易日志已满（50行），需要扩充")
        wb.close()
        return -1

    date_str = datetime.now().strftime("%Y-%m-%d")
    seq = row - 2

    # 安全转换数值
    def _safe_float(v, default=0.0):
        if v is None:
            return default
        try:
            return round(float(v), 6)
        except (TypeError, ValueError):
            return default

    entry_f = _safe_float(entry)
    sl_f = _safe_float(stop_loss)
    tp_f = _safe_float(take_profit)

    # ── 写入数据 ──
    # A=序号 B=信号日 C=币种 D=方向 E=AI入场价 F=AI止损 G=AI止盈
    # H=保证金 I=杠杆 J=名义仓位(公式) K~R=实盘列 S=回测预期 T=偏差 U=备注
    ws.cell(row=row, column=1, value=seq)
    ws.cell(row=row, column=2, value=date_str)
    ws.cell(row=row, column=3, value=symbol)
    ws.cell(row=row, column=4, value=direction)
    ws.cell(row=row, column=5, value=entry_f)
    ws.cell(row=row, column=6, value=sl_f)
    ws.cell(row=row, column=7, value=tp_f)
    ws.cell(row=row, column=8, value=margin)
    ws.cell(row=row, column=9, value=leverage)
    ws.cell(row=row, column=10, value=f"=H{row}*I{row}")
    # K: 实盘入场日 (留空)
    # L: 实盘入场价 (留空)
    # M: 实盘出场日 (留空)
    # N: 实盘出场价 (留空)
    # O: 实盘盈亏($) (公式)
    ws.cell(row=row, column=15, value=(
        f'=IF(OR(L{row}="",N{row}="",D{row}=""),"",'
        f'IF(D{row}="long",(N{row}-L{row})/L{row}*J{row},'
        f'(L{row}-N{row})/L{row}*J{row}))'
    ))
    # P: 实盘盈亏(%)
    ws.cell(row=row, column=16, value=f'=IF(H{row}="","",O{row}/H{row}*100)')
    # Q: 出场原因 (留空)
    # R: 持仓天数
    ws.cell(row=row, column=18, value=f'=IF(OR(K{row}="",M{row}=""),"",DAYS(M{row},K{row}))')
    # S: 回测预期盈亏(%)
    if backtest_expect_pct is not None:
        ws.cell(row=row, column=19, value=round(float(backtest_expect_pct), 2))
    # T: 偏差(%)
    ws.cell(row=row, column=20, value=f'=IF(OR(P{row}="",S{row}=""),"",P{row}-S{row})')
    # U: 备注
    if note:
        ws.cell(row=row, column=21, value=str(note)[:200])

    # ── 格式化 ──
    if direction == "long":
        row_fill = GREEN_FILL
    elif direction == "short":
        row_fill = RED_FILL
    else:
        row_fill = YELLOW_FILL

    for c in range(1, 22):
        cell = ws.cell(row=row, column=c)
        if cell.font == Font():  # 只在默认字体时设置（避免覆盖已有公式的字体）
            pass  # 保留公式的默认字体
        cell.font = Font(name="Arial", size=10)
        cell.alignment = _CENTER
        cell.border = _BORDER if _BORDER else Border()
        if direction in ("long", "short", "neutral"):
            cell.fill = row_fill

    # 数字格式
    for c in [5, 6, 7, 12, 14]:
        ws.cell(row=row, column=c).number_format = '0.000000'
    ws.cell(row=row, column=8).number_format = '0.00'
    ws.cell(row=row, column=10).number_format = '0.00'
    ws.cell(row=row, column=15).number_format = '+0.00;-0.00'
    ws.cell(row=row, column=16).number_format = '0.00'
    ws.cell(row=row, column=19).number_format = '0.00'
    ws.cell(row=row, column=20).number_format = '+0.00;-0.00'

    # 前几列加粗
    for c in [1, 3, 4]:
        ws.cell(row=row, column=c).font = Font(name="Arial", bold=True, size=10)

    try:
        wb.save(LOG_FILE)
        logger.info("信号已写入交易日志: 第%s行 %s %s entry=%s sl=%s tp=%s",
                     row, symbol, direction, entry_f, sl_f, tp_f)
        wb.close()
        return row
    except Exception as e:
        logger.error("保存日志文件失败: %s", e)
        wb.close()
        return -1


def init_log_file() -> bool:
    """如果日志文件不存在，从模板生成一份完整可用的交易日志。"""
    if not _xlsx_available():
        return False

    if os.path.exists(LOG_FILE):
        return True

    logger.info("正在创建交易日志模板: %s", LOG_FILE)

    try:
        wb = Workbook()

        # ═══ Sheet 1: 交易规则 ═══
        ws1 = wb.active
        ws1.title = "交易规则"
        ws1.sheet_properties.tabColor = "16213e"
        ws1.merge_cells("A1:F1")
        ws1["A1"] = "AI策略实盘跟踪表"
        ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="1a1a2e")
        ws1["A1"].alignment = _CENTER
        ws1.row_dimensions[1].height = 30

        ws1.merge_cells("A2:F2")
        ws1["A2"] = "本金 $220  |  杠杆 1x  |  分仓 50/50  |  每笔保证金 $10~15  |  AI自动生成"
        ws1["A2"].font = Font(name="Arial", size=9, color="888888")
        ws1["A2"].alignment = _CENTER

        rules = [
            ("账户参数", ""),
            ("初始入金", "$220"),
            ("杠杆倍数", "1x（资金卫士模式）"),
            ("分仓比例", "50% 首段入场 + 50% 备用金"),
            ("每笔保证金", "$10~$15（本金 5%~7%）"),
            ("名义仓位", "$10~$15（保证金×1x杠杆）"),
            ("单笔最大亏损", "≈$5（止损触发时）"),
            ("最多同时持仓", "1 笔"),
            ("", ""),
            ("纪律铁规", ""),
            ("❶ 每天最多1笔", "与回测逻辑一致"),
            ("❷ 止损绝不手动撤", "触发就认，这是验证系统"),
            ("❸ 超时15天平仓", "对齐回测 max_trade_days=15"),
            ("❹ 不加仓不补仓", "回测没这逻辑，补了无法对比"),
            ("❺ 连亏5笔就停", "实盘与回测差距大，先排查"),
            ("❻ 每笔必填表", "方便和回测报告交叉对比"),
        ]
        row = 4
        for label, val in rules:
            if label in ("账户参数", "纪律铁规"):
                ws1.merge_cells(f"A{row}:F{row}")
                ws1.cell(row=row, column=1, value=f"\u25a0 {label}")
                ws1.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=11, color="e76f51")
                ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FFF3E0")
                row += 1
                continue
            if not label:
                row += 1
                continue
            ws1.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
            ws1.merge_cells(f"B{row}:F{row}")
            ws1.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10, color="333333")
            for c in range(1, 7):
                ws1.cell(row=row, column=c).border = _BORDER
            row += 1

        ws1.column_dimensions["A"].width = 20
        for c in ["B", "C", "D", "E", "F"]:
            ws1.column_dimensions[c].width = 16

        # ═══ Sheet 2: 交易日志 ═══
        ws2 = wb.create_sheet("交易日志")
        ws2.sheet_properties.tabColor = "2a9d8f"
        headers = [
            "序号", "信号日", "币种", "方向", "AI入场价", "AI止损", "AI止盈",
            "保证金($)", "杠杆", "名义仓位($)", "实盘入场日", "实盘入场价",
            "实盘出场日", "实盘出场价", "实盘盈亏($)", "实盘盈亏(%)",
            "出场原因", "持仓天数", "回测预期盈亏(%)", "偏差(%)", "备注"
        ]
        N = len(headers)
        widths = [5, 11, 20, 6, 12, 12, 12, 10, 6, 12, 11, 12, 11, 12, 12, 12, 12, 8, 14, 10, 20]

        ws2.merge_cells(f"A1:{get_column_letter(N)}1")
        ws2["A1"] = "交易日志 — AI 自动填充分析结果"
        ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="1a1a2e")
        ws2["A1"].alignment = _CENTER

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws2.cell(row=2, column=i, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _BORDER
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A3"

        # 预填3~52行
        for r in range(3, 53):
            ws2.cell(row=r, column=1, value=r - 2)
            ws2.cell(row=r, column=10, value=f'=IF(H{r}="","",H{r}*I{r})')
            ws2.cell(row=r, column=15, value=(
                f'=IF(OR(L{r}="",N{r}="",D{r}=""),"",'
                f'IF(D{r}="long",(N{r}-L{r})/L{r}*J{r},'
                f'(L{r}-N{r})/L{r}*J{r}))'
            ))
            ws2.cell(row=r, column=16, value=f'=IF(H{r}="","",O{r}/H{r}*100)')
            ws2.cell(row=r, column=18, value=f'=IF(OR(K{r}="",M{r}=""),"",DAYS(M{r},K{r}))')
            ws2.cell(row=r, column=20, value=f'=IF(OR(P{r}="",S{r}=""),"",P{r}-S{r})')
            for c in range(1, N + 1):
                cell = ws2.cell(row=r, column=c)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = _CENTER
                cell.border = _BORDER
            for c in [5, 6, 7, 12, 14]:
                ws2.cell(row=r, column=c).number_format = '0.000000'
            ws2.cell(row=r, column=8).number_format = '0.00'
            ws2.cell(row=r, column=10).number_format = '0.00'
            ws2.cell(row=r, column=15).number_format = '+0.00;-0.00'
            ws2.cell(row=r, column=16).number_format = '0.00'
            ws2.cell(row=r, column=19).number_format = '0.00'
            ws2.cell(row=r, column=20).number_format = '+0.00;-0.00'

        # 汇总统计
        R = 54
        ws2.merge_cells(f"A{R}:G{R}")
        ws2[f"A{R}"] = "自动汇总统计"
        ws2[f"A{R}"].font = Font(name="Arial", bold=True, size=11, color="1a1a2e")
        ws2[f"A{R}"].fill = PatternFill("solid", fgColor="D6EAF8")

        summary = [
            ("总交易笔数", '=COUNTA(B3:B52)', '0'),
            ("成交笔数", '=COUNTIF(Q3:Q52,"<>未成交")-COUNTIF(Q3:Q52,"")', '0'),
            ("盈利笔数", '=COUNTIF(P3:P52,">0")', '0'),
            ("胜率", '=IF(B55-B57=0,"",D55/(B55-B57))', '0.0%'),
            ("总盈亏($)", '=SUM(O3:O52)', '+$#,##0.00;-$#,##0.00'),
            ("平均盈亏(%)", '=IF(B55=0,"",AVERAGE(P3:P52))', '0.00'),
            ("盈亏比", '=IF(AND(D55>0,E55>0),SUMIF(P3:P52,">0")/ABS(SUMIF(P3:P52,"<0")),"")', '0.00'),
            ("当前资金($)", '=220+SUM(O3:O52)', '$#,##0.00'),
            ("总收益率(%)", '=SUM(O3:O52)/220*100', '0.0%'),
        ]
        for i, (label, formula, fmt) in enumerate(summary):
            r = R + 1 + i
            ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
            ws2.cell(row=r, column=1).border = _BORDER
            ws2.merge_cells(f"B{r}:G{r}")
            ws2.cell(row=r, column=2, value=formula).font = BLUE_FONT
            ws2.cell(row=r, column=2).border = _BORDER
            ws2.cell(row=r, column=2).alignment = _LEFT
            ws2.cell(row=r, column=2).number_format = fmt

        # ═══ Sheet 3: 统计分析 ═══
        ws3 = wb.create_sheet("统计分析")
        ws3.sheet_properties.tabColor = "e76f51"
        ws3.merge_cells("A1:H1")
        ws3["A1"] = "实盘 vs 回测 对比"
        ws3["A1"].font = Font(name="Arial", bold=True, size=13, color="1a1a2e")
        ws3["A1"].alignment = _CENTER

        ws3["A3"] = "\u25a0 整体对比（回测值需手动填入）"
        ws3["A3"].font = Font(name="Arial", bold=True, size=11, color="e76f51")
        comp_h = ["指标", "回测值", "实盘值", "偏差", "判断"]
        for i, h in enumerate(comp_h, 1):
            ws3.cell(row=4, column=i, value=h)
            ws3.cell(row=4, column=i).font = HEADER_FONT
            ws3.cell(row=4, column=i).fill = PatternFill("solid", fgColor="16213e")
            ws3.cell(row=4, column=i).alignment = _CENTER
            ws3.cell(row=4, column=i).border = _BORDER

        comp = [
            ("胜率(%)", "", "='交易日志'!B60", "=C5-B5", '=IF(ABS(D5)<=10,"\u2705 达标","\u26a0 偏差大")'),
            ("平均盈亏(%)", "", "='交易日志'!B62", "=C6-B6", '=IF(ABS(D6)<=5,"\u2705 达标","\u26a0 偏差大")'),
            ("盈亏比", "", "='交易日志'!B65", "=C7-B7", '=IF(C7>=1,"\u2705","\u274c <1")'),
            ("总收益率(%)", "", "='交易日志'!B67", "=C8-B8", '=IF(ABS(D8)<=10,"\u2705 达标","\u26a0 偏差大")'),
        ]
        for i, (label, bt, real, dev, judge) in enumerate(comp):
            r = 5 + i
            ws3.cell(row=r, column=1, value=label)
            ws3.cell(row=r, column=2, value=bt if bt else "（手动填）")
            ws3.cell(row=r, column=3, value=real)
            ws3.cell(row=r, column=4, value=dev)
            ws3.cell(row=r, column=5, value=judge)
            for c in range(1, 6):
                ws3.cell(row=r, column=c).border = _BORDER
                ws3.cell(row=r, column=c).font = Font(name="Arial", size=10)
                ws3.cell(row=r, column=c).alignment = _CENTER
            ws3.cell(row=r, column=2).font = BLUE_FONT

        # 按方向统计
        R3 = 11
        ws3[f"A{R3}"] = "\u25a0 按方向统计"
        ws3[f"A{R3}"].font = Font(name="Arial", bold=True, size=11, color="e76f51")
        dir_h = ["方向", "笔数", "盈利笔数", "胜率", "总盈亏(%)", "均盈亏(%)"]
        for i, h in enumerate(dir_h, 1):
            ws3.cell(row=R3 + 1, column=i, value=h)
            ws3.cell(row=R3 + 1, column=i).font = HEADER_FONT
            ws3.cell(row=R3 + 1, column=i).fill = PatternFill("solid", fgColor="16213e")
            ws3.cell(row=R3 + 1, column=i).alignment = _CENTER
            ws3.cell(row=R3 + 1, column=i).border = _BORDER

        for idx, d in enumerate(["long", "short"]):
            r = R3 + 2 + idx
            ws3.cell(row=r, column=1, value=d).font = Font(name="Arial", bold=True, size=10)
            ws3.cell(row=r, column=2, value=f'=COUNTIF(交易日志!D3:D52,"{d}")')
            ws3.cell(row=r, column=3, value=f'=COUNTIFS(交易日志!D3:D52,"{d}",交易日志!P3:P52,">0")')
            ws3.cell(row=r, column=4, value=f'=IF(B{r}=0,"",C{r}/B{r})')
            ws3.cell(row=r, column=5, value=f'=SUMIF(交易日志!D3:D52,"{d}",交易日志!P3:P52)')
            ws3.cell(row=r, column=6, value=f'=IF(B{r}=0,"",AVERAGEIF(交易日志!D3:D52,"{d}",交易日志!P3:P52))')
            for c in range(1, 7):
                ws3.cell(row=r, column=c).border = _BORDER
                ws3.cell(row=r, column=c).font = Font(name="Arial", size=10)
                ws3.cell(row=r, column=c).alignment = _CENTER
            ws3.cell(row=r, column=4).number_format = '0.0%'
            ws3.cell(row=r, column=6).number_format = '0.00'

        ws3.column_dimensions["A"].width = 14
        for c in ["B", "C", "D", "E", "F"]:
            ws3.column_dimensions[c].width = 16

        # ═══ Sheet 4: 资金曲线 ═══
        ws4 = wb.create_sheet("资金曲线")
        ws4.sheet_properties.tabColor = "264653"
        ws4.merge_cells("A1:D1")
        ws4["A1"] = "资金曲线"
        ws4["A1"].font = Font(name="Arial", bold=True, size=13, color="1a1a2e")
        ws4["A1"].alignment = _CENTER

        curve_h = ["日期", "本笔盈亏($)", "累计盈亏($)", "账户余额($)"]
        for i, h in enumerate(curve_h, 1):
            ws4.cell(row=2, column=i, value=h)
            ws4.cell(row=2, column=i).font = HEADER_FONT
            ws4.cell(row=2, column=i).fill = PatternFill("solid", fgColor="16213e")
            ws4.cell(row=2, column=i).alignment = _CENTER
            ws4.cell(row=2, column=i).border = _BORDER

        ws4.cell(row=3, column=1, value="起始")
        ws4.cell(row=3, column=2, value=0)
        ws4.cell(row=3, column=3, value=0)
        ws4.cell(row=3, column=4, value=220)
        for c in range(1, 5):
            ws4.cell(row=3, column=c).font = Font(name="Arial", size=10)
            ws4.cell(row=3, column=c).alignment = _CENTER
            ws4.cell(row=3, column=c).border = _BORDER

        for r in range(4, 53):
            ws4.cell(row=r, column=1, value="（逐笔填）")
            ws4.cell(row=r, column=3, value=f'=IF(B{r}="","",C{r-1}+B{r})')
            ws4.cell(row=r, column=4, value=f'=220+C{r}')
            for c in range(1, 5):
                ws4.cell(row=r, column=c).font = Font(name="Arial", size=10)
                ws4.cell(row=r, column=c).alignment = _CENTER
                ws4.cell(row=r, column=c).border = _BORDER
            ws4.cell(row=r, column=2).number_format = '+0.00;-0.00'
            ws4.cell(row=r, column=3).number_format = '+0.00;-0.00'
            ws4.cell(row=r, column=4).number_format = '0.00'

        ws4.column_dimensions["A"].width = 14
        for c in ["B", "C", "D"]:
            ws4.column_dimensions[c].width = 16

        wb.save(LOG_FILE)
        wb.close()
        logger.info("交易日志模板已创建: %s", LOG_FILE)
        return True

    except Exception as e:
        logger.exception("创建交易日志模板失败: %s", e)
        return False


def get_write_log_dir() -> str:
    """返回日志文件所在的目录路径。"""
    return os.path.dirname(LOG_FILE)
