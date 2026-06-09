# backtest.py
"""AI 策略日线回测引擎

流程:
  1. 读取 prepare_data.py 预下载的 Parquet 数据
  2. 预计算全部技术指标（无前视偏差）
  3. 遍历每一天 i（warmup 之后），模拟"站在 day i 收盘时"的视角:
     - 计算形态/背离/庄家检测
     - 基于历史资金费率构建情绪数据
     - 调 DeepSeek AI 预测明日走势 + 给出 entry/sl/tp
  4. 模拟交易执行: 从前向 K 线中找入场→止损/止盈触发点
  5. 输出 JSON + HTML 报告

特性:
  - 并发 AI 调用 (ThreadPoolExecutor, 默认4线程)
  - 断点续传 (每批次保存 checkpoint)
  - HTML 报告含权益曲线图 (Chart.js)

用法:
  python backtest.py                          # 回测 config.json 中第一个交易对
  python backtest.py ALLO-USDT-SWAP           # 回测指定交易对
  python backtest.py ALLO-USDT-SWAP 80 150    # 指定起止天数索引
"""

import json
import os
import sys
import time
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

import numpy as np
import pandas as pd

from config import Config
from okx_client import OKXClient
from indicators import calculate_all as calculate_indicators
from patterns import (
    detect_candlestick_patterns,
    detect_rsi_divergence,
    detect_macd_divergence,
    detect_ma_alignment,
)
from manipulation.daily_engine import run_daily_manipulation
from ai_analysis import analyze_daily_with_ai, analyze_with_two_timeframes

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
log = logging.getLogger("backtest")

# ─── 颜色常量 ───
GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
CYAN = "\033[96m"
RESET = "\033[0m"


# ══════════════════════════════════════════════════════════════════
#  回测引擎
# ══════════════════════════════════════════════════════════════════

class BacktestEngine:
    """日线 AI 策略回测引擎。"""

    def __init__(self, cfg: Config, symbol: str):
        self.cfg = cfg
        self.symbol = symbol
        self.safe_name = symbol.replace("-", "_")

        # ── 数据 ──
        self.df: Optional[pd.DataFrame] = None          # K 线
        self.funding_df: Optional[pd.DataFrame] = None  # 资金费率历史

        # ── 结果 ──
        self.trades: List[Dict] = []           # 所有交易（含未成交）
        self.completed_days: Set[int] = set()  # 已分析的天索引

        # ── 可调参数 ──
        self.warmup_days = 65          # 指标预热（MA60=60天 + 余量）
        self.max_trade_days = 15       # 单笔交易最长持仓天数
        self.entry_timeout = 3         # 入场挂单最多等几天
        self.initial_capital = 10000   # 初始资金
        self.position_pct = 0.05       # 每笔交易使用资金比例（避免浮夸复利）
        self.max_workers = 4           # AI 并发数
        self.ai_timeout = 90           # 单次 AI 调用超时(秒)

        # ── 检查点 ──
        self._checkpoint_path: Optional[str] = None
        self._last_ai_direction: str = "neutral"

    # ── 数据加载 ────────────────────────────────────────────

    def load_data(self) -> bool:
        """加载预下载的 Parquet 数据。"""
        hist = self.cfg.HISTORY_DIR
        kline_path = os.path.join(hist, f"kline_{self.safe_name}.csv")
        funding_path = os.path.join(hist, f"funding_{self.safe_name}.csv")

        if not os.path.exists(kline_path):
            log.error(f"K线数据不存在: {kline_path}")
            log.error("请先运行: python prepare_data.py")
            return False

        self.df = pd.read_csv(kline_path, parse_dates=["timestamp"])
        # 去重：按时间戳保序去重（修复旧版本重复写入导致的数据膨胀）
        before = len(self.df)
        self.df = self.df.drop_duplicates(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
        if before != len(self.df):
            log.warning(f"K线去重: {before} → {len(self.df)} (移除 {before - len(self.df)} 条重复)")
        log.info(f"加载K线: {len(self.df)} 根  "
                 f"({self.df['timestamp'].iloc[0]} → {self.df['timestamp'].iloc[-1]})")

        # 数据不足时自动收紧预热要求（至少留 10 根给分析阶段）
        min_needed = self.warmup_days + 10
        if len(self.df) < min_needed:
            new_warmup = max(20, len(self.df) - 10)
            log.warning(f"K线仅 {len(self.df)} 根，warmup 从 {self.warmup_days} → {new_warmup}")
            self.warmup_days = new_warmup

        if os.path.exists(funding_path):
            self.funding_df = pd.read_csv(funding_path, parse_dates=["fundingTime"])
            log.info(f"加载资金费率: {len(self.funding_df)} 条")
        else:
            log.warning("无资金费率缓存，回测中情绪维度将缺失")

        return True

    def _compute_indicators(self):
        """预计算全部技术指标（无前视偏差：每个指标只依赖过去数据）。"""
        log.info("预计算技术指标...")
        self.df = calculate_indicators(self.df, advanced=self.cfg.ADVANCED_INDICATORS)
        log.info("指标计算完成")

    # ── 情绪数据（仅用资金费率历史，无前视偏差）─────────────

    def _get_funding_at(self, cutoff: pd.Timestamp) -> Optional[float]:
        """获取 cutoff 时刻及之前的最近一期资金费率。"""
        if self.funding_df is None or self.funding_df.empty:
            return None
        mask = self.funding_df["fundingTime"] <= cutoff
        if not mask.any():
            return None
        return float(self.funding_df.loc[mask, "fundingRate"].iloc[-1])

    def _build_sentiment(self, day_idx: int) -> Dict:
        """
        为 day_idx 这一天构建情绪数据字典。
        仅使用截止 day_idx 的资金费率历史，无前视偏差。
        OI/多空比在回测中不可得，标记为缺失。
        """
        if self.funding_df is None or self.funding_df.empty:
            return {
                "funding": {"status": "无历史数据"},
                "oi": {"status": "回测不可得"},
                "ls_ratio": {"status": "回测不可得"},
                "overall_bias": "insufficient",
                "summary_text": "回测模式: 情绪数据不可得",
                "warnings": [],
            }

        cutoff = self.df["timestamp"].iloc[day_idx]
        mask = self.funding_df["fundingTime"] <= cutoff
        if not mask.any():
            return {
                "funding": {"status": "无数据"}, "oi": {"status": "回测不可得"},
                "ls_ratio": {"status": "回测不可得"},
                "overall_bias": "insufficient", "summary_text": "", "warnings": [],
            }

        rates = self.funding_df.loc[mask, "fundingRate"].values.astype(float)
        cur = float(rates[-1])
        avg5 = float(np.mean(rates[-5:])) if len(rates) >= 5 else cur
        avg_all = float(np.mean(rates))
        half = max(1, len(rates) // 2)
        trend = "上升" if np.mean(rates[half:]) > np.mean(rates[:half]) else \
                "下降" if np.mean(rates[half:]) < np.mean(rates[:half]) else "持平"

        warnings = []
        if cur > 0.001:
            warnings.append("多头拥挤")
        elif cur < -0.001:
            warnings.append("空头拥挤")

        overall = "bullish" if cur > 0.0005 else "bearish" if cur < -0.0005 else "mixed"

        return {
            "funding": {
                "current": round(cur, 6), "avg_recent_5": round(avg5, 6),
                "avg_all": round(avg_all, 6), "trend": trend,
                "sample_count": len(rates),
                "summary": f"当前{cur:.4%}, 近5期均值{avg5:.4%}, 趋势{trend}",
                "warnings": warnings,
            },
            "oi": {"status": "回测不可得"},
            "ls_ratio": {"status": "回测不可得"},
            "overall_bias": overall,
            "summary_text": f"资金费率: {cur:.4%} ({trend})",
            "warnings": warnings,
        }

    # ── 单日分析（站在 day_idx 收盘时）─────────────────────

    def _analyze_one_day(self, day_idx: int) -> Optional[Dict]:
        """
        模拟站在 day_idx 收盘时，运行完整分析管道 → 调用 AI。
        返回分析结果字典，含 AI 预测。
        """
        try:
            # 只看到 day_idx（含）
            df_view = self.df.iloc[:day_idx + 1].copy()
            n = len(df_view)

            patterns = detect_candlestick_patterns(df_view, lookback=10)
            rsi_div = detect_rsi_divergence(df_view)
            macd_div = detect_macd_divergence(df_view)
            divergence = {"rsi": rsi_div, "macd": macd_div}
            ma_alignment = detect_ma_alignment(df_view)

            manipulation = run_daily_manipulation(
                df_view, symbol=self.symbol,
                wick_shadow_ratio=self.cfg.WICK_SHADOW_RATIO,
            )

            sentiment = self._build_sentiment(day_idx)

            ai_result = analyze_daily_with_ai(
                self.cfg, self.symbol, df_view, n - 1,
                manipulation=manipulation, patterns=patterns,
                divergence=divergence, ma_alignment=ma_alignment,
                sentiment=sentiment,
            )

            return {
                "day_idx": day_idx,
                "date": str(self.df["timestamp"].iloc[day_idx])[:10],
                "close": float(self.df["close"].iloc[day_idx]),
                "ai": ai_result,
            }
        except Exception as e:
            log.error(f"day {day_idx} 分析异常: {e}")
            return None

    # ── 交易模拟（前向遍历）────────────────────────────────

    def _simulate_trade(self, analysis: Dict) -> Optional[Dict]:
        """
        根据 AI 预测模拟交易执行。

        AI 站在 day_idx 收盘时预测"明日"= day_idx+1。
        在前向K线中寻找入场→止损/止盈触发。
        """
        ai = analysis.get("ai", {})
        direction = ai.get("direction", "neutral")
        if direction == "neutral":
            return None

        entry = ai.get("entry")
        stop_loss = ai.get("stop_loss")
        tp1 = ai.get("take_profit1")

        # 安全转 float（AI 可能返回字符串）
        try:
            entry = float(entry) if entry is not None else None
            stop_loss = float(stop_loss) if stop_loss is not None else None
            tp1 = float(tp1) if tp1 is not None else None
        except (ValueError, TypeError):
            return None

        if entry is None or stop_loss is None or entry <= 0 or stop_loss <= 0:
            return None

        day_idx = analysis["day_idx"]
        n_total = len(self.df)
        remaining = n_total - day_idx - 1

        if remaining < 2:
            return None  # 没有足够数据模拟

        # ── 阶段0: 等待入场（最多 entry_timeout 天）─────────
        fill_day = None
        fill_price = None

        for fwd in range(1, min(self.entry_timeout + 1, remaining)):
            k = self.df.iloc[day_idx + fwd]
            lo, hi = float(k["low"]), float(k["high"])
            if lo <= entry <= hi:
                fill_day = day_idx + fwd
                fill_price = entry
                break

        if fill_day is None:
            return {
                "direction": direction,
                "entry": round(entry, 6),
                "stop_loss": round(stop_loss, 6),
                "take_profit": round(tp1, 6),
                "signal_date": str(self.df["timestamp"].iloc[day_idx])[:10],
                "filled": False,
                "exit_reason": "entry_not_reached",
                "pnl_pct": 0.0,
                "r_multiple": 0.0,
                "holding_days": 0,
            }

        # ── 阶段1: 持仓监控 ─────────────────────────────────
        exit_day = None
        exit_price = None
        exit_reason = None
        max_look = min(fill_day + self.max_trade_days, n_total)

        for fwd in range(fill_day + 1, max_look):
            k = self.df.iloc[fwd]
            hi, lo = float(k["high"]), float(k["low"])

            if direction == "long":
                if lo <= stop_loss:
                    exit_day, exit_price, exit_reason = fwd, stop_loss, "stop_loss"
                    break
                if hi >= tp1:
                    exit_day, exit_price, exit_reason = fwd, tp1, "take_profit"
                    break
            else:  # short
                if hi >= stop_loss:
                    exit_day, exit_price, exit_reason = fwd, stop_loss, "stop_loss"
                    break
                if lo <= tp1:
                    exit_day, exit_price, exit_reason = fwd, tp1, "take_profit"
                    break

        # 超时未触发 → 以最后一天收盘价强制平仓
        if exit_day is None:
            exit_day = min(fill_day + self.max_trade_days, n_total - 1)
            exit_price = float(self.df["close"].iloc[exit_day])
            exit_reason = "time_exit"

        # ── 计算盈亏 ────────────────────────────────────────
        if direction == "long":
            pnl_pct = (exit_price - fill_price) / fill_price
        else:
            pnl_pct = (fill_price - exit_price) / fill_price

        risk = abs(fill_price - stop_loss) / fill_price if fill_price > 0 else 0.01
        r_multiple = pnl_pct / risk if risk > 0 else 0.0

        return {
            "direction": direction,
            "entry": round(entry, 6),
            "stop_loss": round(stop_loss, 6),
            "take_profit": round(tp1, 6),
            "signal_date": str(self.df["timestamp"].iloc[day_idx])[:10],
            "fill_date": str(self.df["timestamp"].iloc[fill_day])[:10],
            "fill_price": round(fill_price, 6),
            "exit_date": str(self.df["timestamp"].iloc[exit_day])[:10],
            "exit_price": round(exit_price, 6),
            "exit_reason": exit_reason,
            "filled": True,
            "pnl_pct": round(pnl_pct * 100, 2),
            "r_multiple": round(r_multiple, 2),
            "holding_days": exit_day - fill_day,
        }

    # ── 主循环 ─────────────────────────────────────────────

    def run(self, start_day: int = None, end_day: int = None) -> Dict:
        """执行回测主循环。"""
        # ── 高频模式 ──
        if self.cfg.HIGH_FREQ_MODE:
            log.info("🚀 启用高频模式：AI日线方向 + 1H机械信号")
            if self.df is None:
                self.load_data()
            if self.df is not None:
                self._compute_indicators()
            return self._run_high_freq_backtest()

        if self.df is None and not self.load_data():
            return {"error": "数据加载失败"}

        # ── 预计算全部指标 ──
        self._compute_indicators()

        n = len(self.df)
        actual_start = max(self.warmup_days, start_day or self.warmup_days)
        actual_end = min(n - 5, end_day or n - 5)

        if actual_start >= actual_end:
            return {"error": f"回测范围无效: {actual_start}→{actual_end}, 总K线={n}"}

        log.info(f"回测范围: 第 {actual_start} → {actual_end} 天  "
                 f"({self.df['timestamp'].iloc[actual_start]} → {self.df['timestamp'].iloc[actual_end]})")
        log.info(f"待分析天数: {actual_end - actual_start + 1}, AI并发: {self.max_workers}")

        # ── 加载检查点 ──
        out_dir = self.cfg.OUTPUT_DIR
        os.makedirs(out_dir, exist_ok=True)
        self._checkpoint_path = os.path.join(out_dir, f"bt_ckpt_{self.safe_name}.json")
        self._load_checkpoint()

        # ── 待分析天列表 ──
        pending = [i for i in range(actual_start, actual_end + 1)
                   if i not in self.completed_days]

        if not pending:
            log.info("全部已分析完毕")
            return self._build_report()

        log.info(f"其中 {len(pending)} 天待分析, {len(self.completed_days)} 天已完成")

        # ── 批次并发处理 ──
        batch_size = self.max_workers * 3
        total = len(pending)
        done_count = 0
        t_start = time.time()

        for b_start in range(0, total, batch_size):
            batch = pending[b_start:b_start + batch_size]

            with ThreadPoolExecutor(max_workers=self.max_workers) as pool:
                futures = {pool.submit(self._analyze_one_day, d): d for d in batch}

                for fut in as_completed(futures):
                    day = futures[fut]
                    try:
                        analysis = fut.result(timeout=self.ai_timeout)
                    except Exception as e:
                        log.error(f"day {day} 超时/异常: {e}")
                        self.completed_days.add(day)
                        done_count += 1
                        continue

                    if analysis:
                        trade = self._simulate_trade(analysis)
                        if trade:
                            self.trades.append(trade)
                            if trade.get("filled"):
                                color = GREEN if trade["pnl_pct"] > 0 else RED
                                log.info(f"  [{trade['direction']:>5}] {trade['signal_date']} → "
                                         f"{trade['exit_date']} | "
                                         f"{color}{trade['pnl_pct']:>+7.2f}%{RESET} | "
                                         f"{trade['r_multiple']:>+.2f}R | "
                                         f"{trade['exit_reason']}")
                            else:
                                log.debug(f"  [----] {trade['signal_date']} 未成交")

                    self.completed_days.add(day)
                    done_count += 1

            # 每批次保存检查点
            self._save_checkpoint()

            elapsed = time.time() - t_start
            eta = (elapsed / done_count) * (total - done_count) if done_count > 0 else 0
            log.info(f"进度: {done_count}/{total} ({done_count/total*100:.0f}%)  "
                     f"已耗时{elapsed:.0f}s  预计剩余{eta:.0f}s")

        # ── 最终保存 ──
        self._save_checkpoint()

        # ── 生成当前点位预测（双时间框架：日线+今日1H）──
        log.info("=" * 50)
        log.info("🎯 生成当前交易点位（双时间框架）...")
        self.current_signal = None
        self._current_signal_error = None

        try:
            last_idx = len(self.df) - 2  # 倒数第二根 = 最近完整日线
            min_required = max(15, self.warmup_days)
            if last_idx < min_required:
                msg = (f"数据不足: 需要至少{min_required}天预热(MA/SMA等指标需要)，"
                       f"当前仅{last_idx + 1}根K线 (warmup={self.warmup_days})")
                self._current_signal_error = msg
                log.warning(msg)
            else:
                # Step 1: 计算日线上下文（不调用AI）
                df_view = self.df.iloc[:last_idx + 1].copy()
                patterns = detect_candlestick_patterns(df_view, lookback=10)
                rsi_div = detect_rsi_divergence(df_view)
                macd_div = detect_macd_divergence(df_view)
                divergence = {"rsi": rsi_div, "macd": macd_div}
                ma_alignment = detect_ma_alignment(df_view)
                manipulation = run_daily_manipulation(
                    df_view, symbol=self.symbol,
                    wick_shadow_ratio=self.cfg.WICK_SHADOW_RATIO,
                )
                sentiment = self._build_sentiment(last_idx)

                # Step 2: 拉取今日1H K线
                hourly_df = None
                live_client = OKXClient(self.cfg)
                try:
                    hourly_raw = live_client.get_klines(self.symbol, bar="1H", limit=24)
                    if hourly_raw:
                        hourly_rows = OKXClient.parse_klines(hourly_raw)
                        hourly_df = pd.DataFrame(hourly_rows)
                        hourly_df["timestamp"] = pd.to_datetime(hourly_df["timestamp"], unit="ms")
                        # 只保留今天的（UTC时间）
                        today_utc = pd.Timestamp.utcnow().date()
                        hourly_df["date"] = hourly_df["timestamp"].dt.date
                        hourly_today = hourly_df[hourly_df["date"] == today_utc].sort_values("timestamp")
                        if len(hourly_today) > 0:
                            hourly_df = hourly_today
                            log.info(f"  拉取今日1H K线: {len(hourly_df)} 根")
                        else:
                            hourly_df = hourly_df.sort_values("timestamp").tail(24)
                            log.info(f"  今日无1H数据，使用最近24根1H")
                except Exception as e:
                    log.warning(f"  拉取1H K线失败: {e}，将用纯日线分析")

                # Step 3: 双时间框架AI分析
                ai_result = analyze_with_two_timeframes(
                    self.cfg, self.symbol, self.df, last_idx,
                    hourly_df,
                    manipulation=manipulation, patterns=patterns,
                    divergence=divergence, ma_alignment=ma_alignment,
                    sentiment=sentiment,
                )

                # Step 4: 获取实盘价格（用于前端展示）
                live_price = None
                try:
                    ticker_result = live_client.get_ticker(self.symbol)
                    if ticker_result:
                        live_price = float(ticker_result.get("last", 0))
                except Exception:
                    pass

                # Build current_signal
                self.current_signal = {
                    "day_idx": last_idx,
                    "date": str(self.df["timestamp"].iloc[last_idx])[:10],
                    "close": float(self.df["close"].iloc[last_idx]),
                    "ai": ai_result,
                    "live_price": live_price,
                    "live_time": datetime.now().isoformat(),
                    "_analysis_mode": "dual_timeframe",
                }

                if live_price:
                    log.info(f"  实盘价: {live_price}")

                cs = self.current_signal
                ai = cs.get("ai", {})
                log.info(f"  信号日: {cs['date']}")
                log.info(f"  方向: {ai.get('direction','?')} | 强度: {ai.get('strength','?')}")
                log.info(f"  入场: {ai.get('entry')} | 止损: {ai.get('stop_loss')}")
                log.info(f"  止盈1: {ai.get('take_profit1')} | 止盈2: {ai.get('take_profit2')}")
                log.info(f"  关键支撑: {ai.get('key_support')} | 关键阻力: {ai.get('key_resistance')}")
                log.info(f"  分析逻辑: {ai.get('tomorrow_prediction','?')}")
        except Exception as e:
            self._current_signal_error = f"生成当前信号异常: {e}"
            log.error(f"生成当前信号失败: {e}")

        report = self._build_report()
        return report

    # ══════════════ 高频回测模式 ═══════════════════════════════

    def _run_high_freq_backtest(self) -> Dict:
        """
        高频模式：AI日线定方向 + 1H机械信号 + 固定TP/SL + 移动止损。
        """
        from mechanical_signals import generate_signals_1h, calculate_indicators_1h

        # ── 1. 加载/下载 1H 数据 ──
        hist = self.cfg.HISTORY_DIR
        kline_1h_path = os.path.join(hist, f"kline_1H_{self.safe_name}.csv")

        if not os.path.exists(kline_1h_path):
            log.info("1H 数据不存在，开始下载...")
            client = OKXClient(self.cfg)
            raw = client.get_klines(self.symbol, bar="1H", limit=300)
            if not raw:
                return {"error": "1H 数据下载失败"}
            rows = OKXClient.parse_klines(raw)
            df_1h = pd.DataFrame(rows)
            df_1h["timestamp"] = pd.to_datetime(df_1h["timestamp"], unit="ms")
            df_1h = df_1h.sort_values("timestamp").reset_index(drop=True)
            os.makedirs(hist, exist_ok=True)
            df_1h.to_csv(kline_1h_path, index=False)
            log.info(f"下载1H数据: {len(df_1h)} 根")
        else:
            df_1h = pd.read_csv(kline_1h_path, parse_dates=["timestamp"])
            # 去重
            before = len(df_1h)
            df_1h = df_1h.drop_duplicates(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
            if before != len(df_1h):
                log.warning(f"1H K线去重: {before} → {len(df_1h)}")

        log.info(f"加载1H K线: {len(df_1h)} 根")

        # ── 2. 计算 1H 指标 ──
        log.info("计算1H技术指标...")
        df_1h = calculate_indicators_1h(df_1h)
        log.info(f"1H指标完成: {len(df_1h)} 根")

        # ── 3. 获取日线数据用于 AI 分析 ──
        if self.df is None or self.df.empty:
            from indicators import calculate_all as calculate_indicators
            kline_daily_path = os.path.join(hist, f"kline_{self.safe_name}.csv")
            if not os.path.exists(kline_daily_path):
                return {"error": "日线数据不存在"}
            daily_df = pd.read_csv(kline_daily_path, parse_dates=["timestamp"])
            daily_df = daily_df.drop_duplicates(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
            daily_df = calculate_indicators(daily_df, advanced=self.cfg.ADVANCED_INDICATORS)
        else:
            daily_df = self.df

        # ── 4. 确定日期范围 ──
        df_1h["date"] = df_1h["timestamp"].dt.date
        unique_dates = sorted(df_1h["date"].unique())
        if len(unique_dates) < 30:
            return {"error": f"1H数据天数不足: {len(unique_dates)} 天 (至少需要30天)"}
        trade_dates = unique_dates[30:]  # 1H指标预热需要30根（约1.25天）
        log.info(f"可交易日期: {len(trade_dates)} 天 ({trade_dates[0]} → {trade_dates[-1]})")

        # ── 配置 ──
        tp_pct = self.cfg.TP_PCT
        sl_pct = self.cfg.SL_PCT
        max_sig = self.cfg.MAX_SIGNALS_PER_DAY
        enable_trail = self.cfg.ENABLE_TRAILING_STOP
        trail_act = self.cfg.TRAIL_ACTIVATION_PCT

        # ── 5. 遍历每天 ──
        for day_idx, trade_date in enumerate(trade_dates):
            day_df = df_1h[df_1h["date"] == trade_date].copy()
            if len(day_df) < 3:
                continue

            # 找对应的日线索引（当天或最近的前一个交易日）
            daily_mask = daily_df["timestamp"].dt.date <= trade_date
            if not daily_mask.any():
                continue
            daily_idx = daily_df[daily_mask].index[-1]

            # ── AI 分析（每天一次，只判方向）──
            df_view = daily_df.iloc[:daily_idx + 1].copy()
            n_view = len(df_view)
            if n_view < self.warmup_days:
                continue

            patterns = detect_candlestick_patterns(df_view, lookback=10)
            rsi_div = detect_rsi_divergence(df_view)
            macd_div = detect_macd_divergence(df_view)
            divergence = {"rsi": rsi_div, "macd": macd_div}
            ma_alignment = detect_ma_alignment(df_view)
            manipulation = run_daily_manipulation(
                df_view, symbol=self.symbol,
                wick_shadow_ratio=self.cfg.WICK_SHADOW_RATIO,
            )
            sentiment = self._build_sentiment(daily_idx)

            ai_result = analyze_daily_with_ai(
                self.cfg, self.symbol, df_view, n_view - 1,
                manipulation=manipulation, patterns=patterns,
                divergence=divergence, ma_alignment=ma_alignment,
                sentiment=sentiment,
            )
            ai_dir = ai_result.get("direction", "neutral")
            log.info(f"  [{trade_date}] AI方向: {ai_dir}")

            # ── 生成机械信号 ──
            sigs = generate_signals_1h(day_df, ai_dir, tp_pct, sl_pct, max_sig)
            log.info(f"  [{trade_date}] AI={ai_dir} | 产生 {len(sigs)} 个信号")

            # ── 模拟每笔交易 ──
            for sig in sigs:
                entry_price = sig["entry_price"]
                direction = sig["direction"]
                stop_loss = sig["stop_loss"]
                tp = sig["take_profit"]

                # 找信号所在的小时索引
                sig_ts = sig.get("timestamp")
                sig_idx = None
                for j in range(len(day_df)):
                    r_ts = str(day_df.iloc[j].get("timestamp", ""))
                    if sig_ts and r_ts[:19] == sig_ts:
                        sig_idx = j
                        break
                if sig_idx is None:
                    sig_idx = len(day_df) - 2 if len(day_df) >= 2 else 0

                # ── 入场：下一根 K 线开盘价入场 ──
                fill_idx = sig_idx + 1
                if fill_idx >= len(day_df):
                    continue
                fill_price = float(day_df.iloc[fill_idx]["open"])

                # ── 持仓监控（在当天剩余K线中）──
                exit_idx = None
                exit_price = None
                exit_reason = None
                trailing_sl = stop_loss
                trail_on = False

                for fwd in range(fill_idx + 1, len(day_df)):
                    k = day_df.iloc[fwd]
                    hi = float(k["high"])
                    lo = float(k["low"])
                    cl = float(k["close"])

                    if direction == "long":
                        # 移动止损
                        if enable_trail:
                            profit = (hi - fill_price) / fill_price
                            if profit >= trail_act:
                                trail_on = True
                            if trail_on:
                                lookback_start = max(fill_idx, fwd - 2)
                                recent_low = min(float(day_df["low"].iloc[lookback_start:fwd + 1]))
                                new_sl = recent_low * 0.995  # 留0.5%缓冲
                                trailing_sl = max(trailing_sl, new_sl)

                        if lo <= trailing_sl:
                            exit_idx, exit_price = fwd, trailing_sl
                            exit_reason = "trailing_stop" if trail_on else "stop_loss"
                            break
                        if hi >= tp:
                            exit_idx, exit_price, exit_reason = fwd, tp, "take_profit"
                            break
                    else:  # short
                        if enable_trail:
                            profit = (fill_price - lo) / fill_price
                            if profit >= trail_act:
                                trail_on = True
                            if trail_on:
                                lookback_start = max(fill_idx, fwd - 2)
                                recent_high = max(float(day_df["high"].iloc[lookback_start:fwd + 1]))
                                new_sl = recent_high * 1.005
                                trailing_sl = min(trailing_sl, new_sl)

                        if hi >= trailing_sl:
                            exit_idx, exit_price = fwd, trailing_sl
                            exit_reason = "trailing_stop" if trail_on else "stop_loss"
                            break
                        if lo <= tp:
                            exit_idx, exit_price, exit_reason = fwd, tp, "take_profit"
                            break

                if exit_idx is None:
                    # 当天未出场 → 收盘价平仓
                    exit_idx = len(day_df) - 1
                    exit_price = float(day_df.iloc[exit_idx]["close"])
                    exit_reason = "time_exit"

                # ── 盈亏 ──
                if direction == "long":
                    pnl_pct = (exit_price - fill_price) / fill_price
                else:
                    pnl_pct = (fill_price - exit_price) / fill_price

                risk = abs(fill_price - stop_loss) / fill_price if fill_price > 0 else sl_pct
                r_multiple = pnl_pct / risk if risk > 0 else 0.0

                self.trades.append({
                    "direction": direction,
                    "entry": round(entry_price, 6),
                    "stop_loss": round(stop_loss, 6),
                    "take_profit": round(tp, 6),
                    "signal_date": str(trade_date),
                    "fill_date": str(trade_date),
                    "fill_price": round(fill_price, 6),
                    "exit_date": str(trade_date),
                    "exit_price": round(exit_price, 6),
                    "exit_reason": exit_reason,
                    "filled": True,
                    "pnl_pct": round(pnl_pct * 100, 2),
                    "r_multiple": round(r_multiple, 2),
                    "holding_days": 0,
                    "trigger": sig.get("trigger", ""),
                    "confidence": sig.get("confidence", "medium"),
                })

            # ── 进度 ──
            if (day_idx + 1) % 30 == 0:
                log.info(f"  进度: {day_idx+1}/{len(trade_dates)} 天, "
                         f"信号总数: {len(self.trades)}")

            # ── 检查点保存（每30天）──
            if (day_idx + 1) % 30 == 0:
                self.completed_days.add(day_idx)
                self._save_checkpoint()

        report = self._build_report()
        return report

    # ── 检查点 ─────────────────────────────────────────────

    def _load_checkpoint(self):
        if not self._checkpoint_path or not os.path.exists(self._checkpoint_path):
            return
        try:
            with open(self._checkpoint_path, "r") as f:
                cp = json.load(f)
            cp_days = cp.get("completed_days", [])
            cp_trades = cp.get("trades", [])
            # 验证检查点是否匹配当前数据（天数不能超出K线范围）
            if self.df is not None and cp_days:
                data_len = len(self.df)
                if max(cp_days) >= data_len:
                    log.warning(f"检查点过期(max_day={max(cp_days)} >= data_len={data_len})，丢弃")
                    self._clear_checkpoint()
                    return
            # 交易量不应超过分析天数的1.5倍（每天最多1个信号）
            if len(cp_trades) > len(cp_days) * 1.5:
                log.warning(f"检查点交易量异常({len(cp_trades)}笔 vs {len(cp_days)}天)，丢弃")
                self._clear_checkpoint()
                return
            self.trades = cp_trades
            self.completed_days = set(cp_days)
            log.info(f"从检查点恢复: {len(self.trades)} 笔交易, {len(self.completed_days)} 天已完成")
        except Exception as e:
            log.warning(f"检查点读取失败: {e}")

    def _clear_checkpoint(self):
        """删除过期/损坏的检查点文件。"""
        if self._checkpoint_path and os.path.exists(self._checkpoint_path):
            try:
                os.remove(self._checkpoint_path)
                log.info(f"已删除过期检查点: {self._checkpoint_path}")
            except Exception as e:
                log.warning(f"删除检查点失败: {e}")

    def _save_checkpoint(self):
        if not self._checkpoint_path:
            return
        try:
            cp = {
                "completed_days": sorted(self.completed_days),
                "trades": self.trades,
                "data_length": len(self.df) if self.df is not None else 0,
                "updated": datetime.now().isoformat(),
            }
            with open(self._checkpoint_path, "w") as f:
                json.dump(cp, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            log.warning(f"检查点保存失败: {e}")

    # ══════════════ 报告生成 ═══════════════════════════════

    def _build_report(self) -> Dict:
        """汇总回测结果，生成结构化报告。"""
        filled = [t for t in self.trades if t.get("filled")]
        unfilled = [t for t in self.trades if not t.get("filled")]

        # ── 分离超时平仓（time_exit 不代表 AI 预测正确）──
        signal_trades = [t for t in filled if t.get("exit_reason") != "time_exit"]
        time_exit_trades = [t for t in filled if t.get("exit_reason") == "time_exit"]

        # ── 基础统计（主统计用信号交易，排除超时平仓）──
        stat_base = signal_trades if signal_trades else filled

        wins = [t for t in stat_base if t["pnl_pct"] > 0]
        losses = [t for t in stat_base if t["pnl_pct"] <= 0]

        win_rate = len(wins) / len(stat_base) * 100 if stat_base else 0
        avg_win = np.mean([t["pnl_pct"] for t in wins]) if wins else 0
        avg_loss = np.mean([t["pnl_pct"] for t in losses]) if losses else 0
        avg_r = np.mean([t["r_multiple"] for t in stat_base]) if stat_base else 0

        total_win = sum(t["pnl_pct"] for t in wins)
        total_loss = abs(sum(t["pnl_pct"] for t in losses))
        profit_factor = total_win / total_loss if total_loss > 0 else (float("inf") if total_win > 0 else 0)

        # ── 连续盈亏（仅信号交易）──
        streak_win = streak_loss = max_streak_win = max_streak_loss = 0
        for t in stat_base:
            if t["pnl_pct"] > 0:
                streak_win += 1; streak_loss = 0
                max_streak_win = max(max_streak_win, streak_win)
            else:
                streak_loss += 1; streak_win = 0
                max_streak_loss = max(max_streak_loss, streak_loss)

        # ── 权益曲线 ──
        equity = self.initial_capital
        equity_curve = []
        peak = equity
        max_dd = 0.0

        sorted_trades = sorted(
            [t for t in filled if t.get("fill_date")],
            key=lambda x: x["fill_date"]
        )

        for t in sorted_trades:
            equity *= (1 + t["pnl_pct"] / 100 * self.position_pct)
            equity_curve.append({
                "date": t["exit_date"],
                "equity": round(equity, 2),
                "pnl_pct": t["pnl_pct"],
            })
            peak = max(peak, equity)
            dd = (peak - equity) / peak if peak > 0 else 0
            max_dd = max(max_dd, dd)

        total_return = (equity - self.initial_capital) / self.initial_capital * 100 if filled else 0
        # 简单累加总收益（非复利，更直观）
        simple_total_return = sum(t["pnl_pct"] for t in filled) if filled else 0

        # ── 夏普比率 ──
        rets = [t["pnl_pct"] / 100 for t in filled]
        sharpe = (np.mean(rets) / np.std(rets) * np.sqrt(252)) if len(rets) > 1 and np.std(rets) > 0 else 0

        # ── 按月统计 ──
        monthly = {}
        for t in sorted_trades:
            m = t["exit_date"][:7]  # YYYY-MM
            monthly.setdefault(m, []).append(t["pnl_pct"])
        monthly_summary = {
            m: {
                "count": len(v),
                "total_pnl": round(sum(v), 2),
                "win_rate": round(len([x for x in v if x > 0]) / len(v) * 100, 1),
            }
            for m, v in sorted(monthly.items())
        }

        # ── 按方向 ──
        by_dir = {}
        for d in ["long", "short"]:
            subset = [t for t in filled if t["direction"] == d]
            if subset:
                w = len([t for t in subset if t["pnl_pct"] > 0])
                by_dir[d] = {
                    "count": len(subset), "win_rate": round(w / len(subset) * 100, 1),
                    "avg_pnl": round(np.mean([t["pnl_pct"] for t in subset]), 2),
                    "avg_r": round(np.mean([t["r_multiple"] for t in subset]), 2),
                }

        # ── 按出场原因 ──
        by_reason = {}
        for r in set(t["exit_reason"] for t in filled):
            subset = [t for t in filled if t["exit_reason"] == r]
            w = len([t for t in subset if t["pnl_pct"] > 0])
            by_reason[r] = {
                "count": len(subset),
                "win_rate": round(w / len(subset) * 100, 1) if subset else 0,
                "avg_pnl": round(np.mean([t["pnl_pct"] for t in subset]), 2),
            }

        period_start = str(self.df["timestamp"].iloc[self.warmup_days])[:10]
        period_end = str(self.df["timestamp"].iloc[-1])[:10]

        return {
            "symbol": self.symbol,
            "period": {"start": period_start, "end": period_end},
            "current_signal": getattr(self, "current_signal", None),
            "current_signal_error": getattr(self, "_current_signal_error", None),
            "summary": {
                "total_signals": len(self.trades),
                "filled_trades": len(filled),
                "unfilled_trades": len(unfilled),
                "signal_trades": len(signal_trades),        # 非超时成交
                "time_exit_trades": len(time_exit_trades),   # 超时平仓
                "win_rate": round(win_rate, 1),
                "avg_win_pct": round(avg_win, 2),
                "avg_loss_pct": round(avg_loss, 2),
                "avg_r_multiple": round(avg_r, 2),
                "profit_factor": round(profit_factor, 2),
                "total_return_pct": round(total_return, 2),
                "simple_total_return_pct": round(simple_total_return, 2),
                "max_drawdown_pct": round(max_dd * 100, 2),
                "sharpe_ratio": round(sharpe, 2),
                "max_streak_win": max_streak_win,
                "max_streak_loss": max_streak_loss,
                "initial_capital": self.initial_capital,
                "final_capital": round(equity, 2),
                "position_pct": int(self.position_pct * 100),
            },
            "by_direction": by_dir,
            "by_exit_reason": by_reason,
            "monthly": monthly_summary,
            "trades": self.trades,  # 全部交易记录（含已成交和未成交）
            "equity_curve": equity_curve,
        }


# ══════════════════════════════════════════════════════════════════
#  输出
# ══════════════════════════════════════════════════════════════════

def print_report(report: Dict):
    """控制台打印回测报告。"""
    s = report.get("summary", {})

    # ── 当前点位（放在最上面）──
    cs = report.get("current_signal")
    if cs:
        ai = cs.get("ai", {})
        direction = ai.get("direction", "neutral")
        dir_emoji = {"long": "🟢做多", "short": "🔴做空", "neutral": "⚪观望"}.get(direction, "⚪观望")
        print("\n" + "=" * 70)
        print("  🎯 当 前 交 易 点 位（基于最新日线 AI 预测）")
        print("=" * 70)
        print(f"  信号日     : {cs.get('date','?')}")
        print(f"  收盘价     : {cs.get('close','?')}")
        print(f"  AI方向     : {dir_emoji} | 强度: {ai.get('strength','?')}")
        print(f"  {'─'*50}")
        print(f"  入场挂单   : {ai.get('entry') or 'N/A'}")
        print(f"  止损价     : {ai.get('stop_loss') or 'N/A'}")
        print(f"  止盈1      : {ai.get('take_profit1') or 'N/A'}")
        print(f"  止盈2      : {ai.get('take_profit2') or 'N/A'}")
        print(f"  关键支撑   : {ai.get('key_support') or 'N/A'}")
        print(f"  关键阻力   : {ai.get('key_resistance') or 'N/A'}")
        print(f"  明日预测   : {ai.get('tomorrow_prediction') or 'N/A'}")
        print("=" * 70)

    print("\n" + "=" * 70)
    print("  📊 回 测 报 告")
    print("=" * 70)
    print(f"  交易对     : {report.get('symbol', '?')}")
    print(f"  回测周期   : {report.get('period', {}).get('start', '?')}"
          f" → {report.get('period', {}).get('end', '?')}")
    print(f"  总信号/成交/未成交 : {s.get('total_signals', 0)} / "
          f"{s.get('filled_trades', 0)} / {s.get('unfilled_trades', 0)}")
    print(f"  （其中信号交易 : {s.get('signal_trades', 0)} 笔, "
          f"超时平仓 : {s.get('time_exit_trades', 0)} 笔）")
    print(f"  {'─'*50}")
    print(f"  胜率       : {s.get('win_rate', 0):.1f}%  (仅统计信号交易)")
    print(f"  平均盈利   : {s.get('avg_win_pct', 0):.2f}%")
    print(f"  平均亏损   : {s.get('avg_loss_pct', 0):.2f}%")
    print(f"  盈亏比     : {s.get('profit_factor', 0):.2f}")
    print(f"  平均R倍数  : {s.get('avg_r_multiple', 0):.2f}R")
    print(f"  最长连胜   : {s.get('max_streak_win', 0)} 笔")
    print(f"  最长连亏   : {s.get('max_streak_loss', 0)} 笔")
    print(f"  {'─'*50}")
    print(f"  简单累加收益 : {s.get('simple_total_return_pct', 0):+.2f}%  (所有成交盈亏累加)")
    print(f"  复利收益率   : {s.get('total_return_pct', 0):+.2f}%  ({s.get('position_pct',5)}%仓位复利)")
    print(f"  最大回撤   : {s.get('max_drawdown_pct', 0):.2f}%")
    print(f"  夏普比率   : {s.get('sharpe_ratio', 0):.2f}")
    print(f"  初始资金   : ${s.get('initial_capital', 0):,.0f}")
    print(f"  最终资金   : ${s.get('final_capital', 0):,.2f}")

    # 按方向
    by_dir = report.get("by_direction", {})
    if by_dir:
        print(f"\n  按方向:")
        for d, v in by_dir.items():
            print(f"    {d:<6} {v['count']}笔  胜率{v['win_rate']}%  "
                  f"均盈亏{v['avg_pnl']:+.2f}%  R={v['avg_r']:.2f}")

    # 按月
    monthly = report.get("monthly", {})
    if monthly:
        print(f"\n  按月:")
        for m, v in monthly.items():
            print(f"    {m}  {v['count']}笔  总盈亏{v['total_pnl']:+.2f}%  胜率{v['win_rate']}%")

    # 最近交易
    trades = report.get("trades", [])
    if trades:
        print(f"\n  最近10笔交易:")
        print(f"  {'信号日':<12} {'出场日':<12} {'方向':<6} {'盈亏':>8} {'R':>6} {'原因'}")
        print(f"  {'-'*55}")
        for t in trades[-10:]:
            color = GREEN if t["pnl_pct"] > 0 else RED
            print(f"  {t['signal_date']:<12} {t['exit_date']:<12} {t['direction']:<6} "
                  f"{color}{t['pnl_pct']:>+7.2f}%{RESET} {t['r_multiple']:>+5.2f}R "
                  f"{t['exit_reason']}")

    print("=" * 70)


def save_report(report: Dict, cfg: Config):
    """保存 JSON、HTML 和 Excel 报告（回测明细 + 统计 + 实盘跟踪）。"""
    out_dir = cfg.OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    symbol = report.get("symbol", "UNKNOWN").replace("-", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # ── JSON ──
    json_path = os.path.join(out_dir, f"backtest_{symbol}_{ts}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n📄 JSON 报告: {json_path}")

    # ── HTML ──
    html = _build_html(report)
    html_path = os.path.join(out_dir, f"backtest_{symbol}_{ts}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"🌐 HTML 报告: {html_path}")

    # ── Excel: 回测明细 + 回测统计 + 实盘跟踪 ──
    _save_excel(report, cfg, ts)


def _save_excel(report: Dict, cfg: Config, ts: str):
    """将回测交易 + AI分析信号汇总写入 AI策略交易日志.xlsx 的"交易日志" Sheet，
       用"来源"列区分"回测"与"分析"。回测统计单独保留一个 Sheet。"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Warning: openpyxl not installed, skip Excel. pip install openpyxl")
        return

    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AI策略交易日志.xlsx")
    symbol = report.get("symbol", "UNKNOWN")

    if os.path.exists(log_path):
        wb = load_workbook(log_path)
        print(f"[Excel] 加载: {log_path}")
    else:
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "交易日志"
        print(f"[Excel] 新建: {log_path}")

    # ── 样式 ──
    hdr_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tbl_hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    tbl_hdr_font = Font(name="Arial", size=10, bold=True)
    dfont = Font(name="Arial", size=10)
    green_f = Font(name="Arial", size=10, color="006100")
    green_bg = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_f = Font(name="Arial", size=10, color="9C0006")
    red_bg = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    src_fill_bt = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 回测=浅黄
    src_fill_ai = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # 分析=浅绿
    bdr = Border(left=Side("thin","B0B0B0"), right=Side("thin","B0B0B0"),
                 top=Side("thin","B0B0B0"), bottom=Side("thin","B0B0B0"))
    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left", vertical="center")
    ttl_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    sub_font = Font(name="Arial", size=12, bold=True, color="2B579A")

    def pnl_style(cell):
        try:
            v = float(cell.value) if cell.value is not None else 0
            if v > 0: cell.font = green_f; cell.fill = green_bg
            elif v < 0: cell.font = red_f; cell.fill = red_bg
        except: pass

    # ══════════════════════════════════════════════════════════════
    #  交易日志 Sheet — 统一汇总
    # ══════════════════════════════════════════════════════════════
    if "交易日志" not in wb.sheetnames:
        tw = wb.create_sheet("交易日志")
    else:
        tw = wb["交易日志"]

    SOURCE_COL = 22  # "来源"列

    # ── 确保表头完整（含"来源"列）──
    headers = ["序号","信号日","币种","方向","AI入场价","AI止损","AI止盈",
               "保证金($)","杠杆","名义仓位($)","实盘入场日","实盘入场价",
               "实盘出场日","实盘出场价","实盘盈亏($)","实盘盈亏(%)",
               "出场原因","持仓天数","回测预期盈亏(%)","偏差(%)","备注","来源"]

    # 检测是否已有"来源"列，没有则补齐
    existing_src = tw.cell(row=1, column=SOURCE_COL).value
    if existing_src != "来源" and tw.cell(row=2, column=SOURCE_COL).value != "来源":
        # 找表头行（可能是 row 1 或 row 2）
        hdr_row_num = None
        for chk in [1, 2]:
            if tw.cell(row=chk, column=1).value in (None, "序号"):
                hdr_row_num = chk
        if hdr_row_num is None:
            hdr_row_num = 2
        # 写入完整表头
        for c, h in enumerate(headers, 1):
            cell = tw.cell(row=hdr_row_num, column=c)
            cell.value = h
            cell.font = tbl_hdr_font
            cell.fill = tbl_hdr_fill
            cell.alignment = ca
            cell.border = bdr

    # ── 读取现有分析信号（手动填的，signal_date 非空 且 来源不为"回测"）──
    data_start = 2  # 表头在第 2 行
    existing_analysis = []
    for ri in range(data_start + 1, tw.max_row + 1):
        sd = tw.cell(row=ri, column=2).value
        src = tw.cell(row=ri, column=SOURCE_COL).value
        if sd and src != "回测":
            row_data = {}
            for c in range(1, SOURCE_COL + 1):
                row_data[c] = tw.cell(row=ri, column=c).value
            existing_analysis.append(row_data)

    # ── 清空所有数据行（保留表头行 1-2）──
    for ri in range(tw.max_row, data_start, -1):
        tw.delete_rows(ri)

    # ── 格式常量 ──
    reason_cn = {"stop_loss":"止损","take_profit":"止盈","time_exit":"超时",
                 "trailing_stop":"移动止损","entry_not_reached":"未成交"}
    REASON_COL = 17; PNL_COL = 16; EXPECT_COL = 19; DEVIATION_COL = 20; NOTE_COL = 21

    def _write_row_base(ws, row, col_values):
        """写入一行数据并设置统一样式，col_values = {col_num: value}"""
        for c in range(1, SOURCE_COL + 1):
            cell = ws.cell(row=row, column=c)
            if c in col_values:
                cell.value = col_values[c]
            elif c not in col_values:
                # 不在指定列中的，清空（避免旧数据残留）
                if c not in [col_values.get('_skip_clear')]:  # 已显式设置的不清
                    pass  # 下方显式清空
            cell.font = dfont
            cell.alignment = ca if c > 1 else la
            cell.border = bdr
        # 显式清空未设置的列，避免残留
        for c in range(1, SOURCE_COL + 1):
            if c not in col_values:
                ws.cell(row=row, column=c).value = None
        # PNL 着色
        if PNL_COL in col_values:
            pnl_style(ws.cell(row=row, column=PNL_COL))

    # ═══ 1. 写入回测交易（先已成交，再未成交） ═══
    all_trades = sorted(report.get("trades", []), key=lambda x: x.get("signal_date", ""))
    filled_trades = [t for t in all_trades if t.get("filled")]
    unfilled_trades = [t for t in all_trades if not t.get("filled")]
    write_row = data_start + 1
    bt_count = 0

    # ── 1a. 已成交交易：完整写入 ──
    for t in filled_trades:
        d = t.get("direction","")
        e = t.get("entry", 0)
        sl = t.get("stop_loss", 0)
        tp = t.get("take_profit", 0)
        pnl = t.get("pnl_pct", 0)
        fp = t.get("fill_price", 0)
        ep_val = t.get("exit_price", 0)

        vals = {
            2: t.get("signal_date",""),
            3: symbol,
            4: d,
            5: round(e, 6) if e else None,
            6: round(sl, 6) if sl else None,
            7: round(tp, 6) if tp else None,
            8: 10, 9: 2,
            11: t.get("fill_date",""),
            12: round(fp, 6) if fp else None,
            13: t.get("exit_date",""),
            14: round(ep_val, 6) if ep_val else None,
            PNL_COL: round(pnl, 2),
            REASON_COL: reason_cn.get(t.get("exit_reason",""), t.get("exit_reason","")),
            18: t.get("holding_days", 0),
            SOURCE_COL: "回测",
        }
        if e and tp and e > 0:
            vals[EXPECT_COL] = round((tp - e) / e * 100 if d == "long" else (e - tp) / e * 100, 2)

        _write_row_base(tw, write_row, vals)
        tw.cell(row=write_row, column=SOURCE_COL).fill = src_fill_bt
        tw.cell(row=write_row, column=PNL_COL).value = vals.get(PNL_COL)
        pnl_style(tw.cell(row=write_row, column=PNL_COL))
        write_row += 1
        bt_count += 1

    # ── 1b. 未成交交易：写入信号点位，出场原因填"未成交" ──
    unfilled_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 浅橙色底
    for t in unfilled_trades:
        d = t.get("direction","")
        e = t.get("entry", 0)
        sl = t.get("stop_loss", 0)
        tp = t.get("take_profit", 0)

        vals = {
            2: t.get("signal_date",""),
            3: symbol,
            4: d,
            5: round(e, 6) if e else None,
            6: round(sl, 6) if sl else None,
            7: round(tp, 6) if tp else None,
            8: 10, 9: 2,
            11: "", 12: None, 13: "", 14: None,  # 实盘列留空
            PNL_COL: None,
            REASON_COL: "未成交",
            18: 0,
            SOURCE_COL: "回测",
        }
        if e and tp and e > 0:
            vals[EXPECT_COL] = round((tp - e) / e * 100 if d == "long" else (e - tp) / e * 100, 2)

        _write_row_base(tw, write_row, vals)
        tw.cell(row=write_row, column=SOURCE_COL).fill = unfilled_fill
        write_row += 1
        bt_count += 1

    # ═══ 2. 补写原有的分析信号 ═══
    for row_data in existing_analysis:
        _write_row_base(tw, write_row, {c: row_data[c] for c in range(1, SOURCE_COL)})
        tw.cell(row=write_row, column=SOURCE_COL).value = "分析"
        tw.cell(row=write_row, column=SOURCE_COL).font = dfont
        tw.cell(row=write_row, column=SOURCE_COL).fill = src_fill_ai
        tw.cell(row=write_row, column=SOURCE_COL).border = bdr
        write_row += 1

    # ═══ 3. 写入当前 AI 信号 ═══
    cs = report.get("current_signal")
    if cs:
        ai = cs.get("ai", {})
        direction = ai.get("direction", "neutral")
        if direction != "neutral":
            entry = ai.get("entry")
            sl = ai.get("stop_loss")
            tp = ai.get("take_profit1")
            vals = {
                2: cs.get("date",""), 3: symbol, 4: direction,
                8: 10, 9: 2, SOURCE_COL: "分析",
            }
            if entry: vals[5] = round(entry, 6)
            if sl: vals[6] = round(sl, 6)
            if tp: vals[7] = round(tp, 6)
            if entry and tp and entry > 0:
                exp = (tp - entry) / entry * 100 if direction == "long" else (entry - tp) / entry * 100
                vals[EXPECT_COL] = round(exp, 2)
            _write_row_base(tw, write_row, vals)
            tw.cell(row=write_row, column=SOURCE_COL).fill = src_fill_ai
            print(f"  [Log] 交易日志已追加AI信号: {direction} {entry}")
            write_row += 1

    # ═══ 4. 统一编号 ═══
    seq = 1
    for ri in range(data_start + 1, write_row):
        if tw.cell(row=ri, column=2).value:
            tw.cell(row=ri, column=1).value = seq
            seq += 1

    # 列宽
    widths = {1:5, 2:12, 3:16, 4:6, 5:12, 6:12, 7:12, 8:9, 9:5, 10:10,
              11:12, 12:12, 13:12, 14:12, 15:12, 16:10, 17:8, 18:8, 19:12, 20:8, 21:30, SOURCE_COL:8}
    for c, w in widths.items():
        tw.column_dimensions[get_column_letter(c)].width = w

    # ══════════════════════════════════════════════════════════════
    #  回测统计 Sheet（汇总数据）
    # ══════════════════════════════════════════════════════════════
    stat_name = "回测统计"
    if stat_name in wb.sheetnames:
        del wb[stat_name]
    ws2 = wb.create_sheet(stat_name)
    token = datetime.now().strftime("%Y%m%d_%H%M")
    s = report.get("summary", {})
    period = report.get("period", {})

    ws2.merge_cells("A1:D1")
    ws2.cell(row=1, column=1, value=f"回测统计 - {symbol}  [{token}]").font = ttl_font
    ws2.merge_cells("A2:D2")
    ws2.cell(row=2, column=1, value=f"周期: {period.get('start','?')} -> {period.get('end','?')}").font = sub_font

    r = 4
    for sec_title, rows_data in [
        ("核心指标", [
            ("总信号 / 成交 / 未成交", f"{s.get('total_signals',0)} / {s.get('filled_trades',0)} / {s.get('unfilled_trades',0)}"),
            ("信号交易(非超时) / 超时平仓", f"{s.get('signal_trades',0)} / {s.get('time_exit_trades',0)}"),
            ("胜率(信号交易)", f"{s.get('win_rate',0):.1f}%"),
            ("平均盈利 / 平均亏损", f"{s.get('avg_win_pct',0):.2f}% / {s.get('avg_loss_pct',0):.2f}%"),
            ("盈亏比", f"{s.get('profit_factor',0):.2f}"),
            ("平均R倍数", f"{s.get('avg_r_multiple',0):.2f}R"),
            (f"复利收益({s.get('position_pct',5)}%仓位) / 简单累加", f"{s.get('total_return_pct',0):+.2f}% / {s.get('simple_total_return_pct',0):+.2f}%"),
            ("最大回撤", f"{s.get('max_drawdown_pct',0):.2f}%"),
            ("夏普比率", f"{s.get('sharpe_ratio',0):.2f}"),
            ("最长连胜 / 连亏", f"{s.get('max_streak_win',0)} 笔 / {s.get('max_streak_loss',0)} 笔"),
            ("初始资金 -> 最终资金", f"${s.get('initial_capital',0):,.0f} -> ${s.get('final_capital',0):,.2f}"),
        ]),
        ("按月统计", [([m, v["count"], v["total_pnl"], f"{v['win_rate']}%"], 3)
                      for m, v in sorted(report.get("monthly", {}).items())]),
        ("按方向统计", [([d, v["count"], f"{v['win_rate']}%", v["avg_pnl"], v["avg_r"]], 4)
                       for d, v in report.get("by_direction", {}).items()]),
        ("按出场原因", [([{"stop_loss":"止损","take_profit":"止盈","time_exit":"超时","trailing_stop":"移动止损"}.get(k,k),
                          v["count"], f"{v['win_rate']}%", v["avg_pnl"]], 4)
                        for k, v in sorted(report.get("by_exit_reason", {}).items())]),
    ]:
        if not rows_data:
            continue
        ws2.merge_cells(f"A{r}:D{r}")
        ws2.cell(row=r, column=1, value=sec_title).font = sub_font
        r += 1
        if sec_title == "核心指标":
            for label, value in rows_data:
                ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
                ws2.cell(row=r, column=1).border = bdr; ws2.cell(row=r, column=1).alignment = la
                ws2.merge_cells(f"B{r}:D{r}")
                ws2.cell(row=r, column=2, value=value).font = dfont
                ws2.cell(row=r, column=2).border = bdr
                r += 1
        else:
            # 表格式区块
            if sec_title == "按月统计":
                hdrs = ["月份","笔数","总盈亏(%)","胜率"]
            elif sec_title == "按方向统计":
                hdrs = ["方向","笔数","胜率","平均盈亏(%)","平均R倍数"]
            else:
                hdrs = ["出场原因","笔数","胜率","平均盈亏(%)"]
            for ci, h in enumerate(hdrs, 1):
                cell = ws2.cell(row=r, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = bdr
            r += 1
            for vals, pnl_col in rows_data:
                for ci, v in enumerate(vals, 1):
                    cell = ws2.cell(row=r, column=ci, value=v)
                    cell.font = dfont; cell.alignment = ca if ci > 1 else la; cell.border = bdr
                    if ci == pnl_col:
                        try:
                            fv = float(v) if isinstance(v, (int,float)) else float(str(v).replace("%",""))
                            if fv > 0: cell.font = green_f; cell.fill = green_bg
                            elif fv < 0: cell.font = red_f; cell.fill = red_bg
                        except: pass
                r += 1
        r += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18

    # ══════════════════════════════════════════════════════════════
    #  保存
    # ══════════════════════════════════════════════════════════════
    wb.save(log_path)
    print(f"[Excel] 已写入: {log_path}  (回测{bt_count}笔 + 分析信号)")


def _build_html(report: Dict) -> str:
    """生成带 Chart.js 权益曲线的 HTML 报告。"""
    s = report.get("summary", {})
    eq = report.get("equity_curve", [])
    trades = report.get("trades", [])
    cs = report.get("current_signal")

    # ── 当前信号卡片 ──
    signal_html = ""
    if cs:
        ai = cs.get("ai", {})
        direction = ai.get("direction", "neutral")
        dir_color = {"long": "#22c55e", "short": "#ef4444", "neutral": "#8b949e"}.get(direction, "#8b949e")
        dir_text = {"long": "🟢 做多", "short": "🔴 做空", "neutral": "⚪ 观望"}.get(direction, "⚪ 观望")
        signal_html = f"""<div class="chart-box" style="border-color:{dir_color};border-width:2px;">
<h3>🎯 当前交易点位（AI预测） <span style="color:{dir_color}">{dir_text}</span> · 强度{ai.get('strength','?')}</h3>
<div class="stats" style="margin-top:12px">
<div class="stat-card"><div class="stat-value">{cs.get('date','?')}</div><div class="stat-label">信号日</div></div>
<div class="stat-card"><div class="stat-value">{cs.get('close','?')}</div><div class="stat-label">收盘价</div></div>
<div class="stat-card"><div class="stat-value" style="color:#58a6ff">{ai.get('entry','N/A')}</div><div class="stat-label">入场挂单</div></div>
<div class="stat-card"><div class="stat-value" style="color:#ef4444">{ai.get('stop_loss','N/A')}</div><div class="stat-label">止损价</div></div>
<div class="stat-card"><div class="stat-value" style="color:#22c55e">{ai.get('take_profit1','N/A')}</div><div class="stat-label">止盈1</div></div>
<div class="stat-card"><div class="stat-value" style="color:#22c55e">{ai.get('take_profit2','N/A')}</div><div class="stat-label">止盈2</div></div>
</div>
<div style="margin-top:10px;display:grid;grid-template-columns:1fr 1fr;gap:8px">
<div style="background:#1c2129;padding:8px 12px;border-radius:6px;font-size:13px">
<span style="color:#8b949e">关键支撑:</span> <strong>{ai.get('key_support','N/A')}</strong></div>
<div style="background:#1c2129;padding:8px 12px;border-radius:6px;font-size:13px">
<span style="color:#8b949e">关键阻力:</span> <strong>{ai.get('key_resistance','N/A')}</strong></div>
</div>
<p style="margin-top:8px;color:#8b949e;font-size:13px">📝 {ai.get('tomorrow_prediction','N/A')}</p>
</div>"""

    eq_labels = json.dumps([e["date"] for e in eq])
    eq_values = json.dumps([e["equity"] for e in eq])

    # 交易明细行
    rows = ""
    for t in trades:
        c = "#22c55e" if t["pnl_pct"] > 0 else "#ef4444"
        reason_cn = {"stop_loss": "止损", "take_profit": "止盈", "time_exit": "超时"}.get(
            t.get("exit_reason", ""), t.get("exit_reason", "")
        )
        rows += f"""<tr>
            <td>{t.get('signal_date','')}</td><td>{t.get('fill_date','')}</td>
            <td>{t.get('exit_date','')}</td><td>{t.get('direction','')}</td>
            <td>{t.get('fill_price',0):.6f}</td><td>{t.get('exit_price',0):.6f}</td>
            <td style="color:#ef4444">{t.get('stop_loss',0):.6f}</td>
            <td style="color:#22c55e">{t.get('take_profit',0):.6f}</td>
            <td style="color:{c};font-weight:700">{t.get('pnl_pct',0):+.2f}%</td>
            <td>{t.get('r_multiple',0):+.2f}R</td>
            <td>{t.get('holding_days',0)}天</td>
            <td>{reason_cn}</td>
        </tr>"""

    # 按月统计
    monthly_rows = ""
    for m, v in report.get("monthly", {}).items():
        c = "#22c55e" if v["total_pnl"] > 0 else "#ef4444"
        monthly_rows += f"""<tr>
            <td>{m}</td><td>{v['count']}</td>
            <td style="color:{c};font-weight:700">{v['total_pnl']:+.2f}%</td>
            <td>{v['win_rate']}%</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>回测报告 - {report.get('symbol','')}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d1117;color:#c9d1d9;padding:20px}}
h1{{text-align:center;margin-bottom:4px;color:#58a6ff}}
.container{{max-width:1200px;margin:0 auto}}
.stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-bottom:20px}}
.stat-card{{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:14px;text-align:center}}
.stat-value{{font-size:26px;font-weight:700;color:#58a6ff}}
.stat-value.up{{color:#22c55e}}.stat-value.down{{color:#ef4444}}
.stat-label{{font-size:12px;color:#8b949e;margin-top:4px}}
.chart-box,.table-box{{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:18px;margin-bottom:20px}}
table{{width:100%;border-collapse:collapse}}
th,td{{padding:8px 12px;text-align:left;border-bottom:1px solid #30363d;font-size:13px}}
th{{background:#21262d;color:#8b949e;font-weight:600;position:sticky;top:0}}
tr:hover{{background:#1c2129}}
.footer{{text-align:center;color:#484f58;font-size:11px;margin-top:24px}}
h3{{margin-bottom:12px;color:#c9d1d9}}
</style>
</head>
<body>
<div class="container">
<h1>📊 回测报告</h1>
<p style="text-align:center;color:#8b949e;margin-bottom:16px">
{report.get('symbol','')} &nbsp;|&nbsp;
{report.get('period',{}).get('start','?')} → {report.get('period',{}).get('end','?')}
</p>

<div class="stats">
<div class="stat-card"><div class="stat-value {'up' if s.get('simple_total_return_pct',0)>=0 else 'down'}">{s.get('simple_total_return_pct',0):+.1f}%</div><div class="stat-label">简单累加收益 (所有成交)</div></div>
<div class="stat-card"><div class="stat-value {'up' if s.get('total_return_pct',0)>=0 else 'down'}">{s.get('total_return_pct',0):+.1f}%</div><div class="stat-label">复利收益 ({s.get('position_pct',5)}%仓位)</div></div>
<div class="stat-card"><div class="stat-value">{s.get('win_rate',0):.1f}%</div><div class="stat-label">胜率 (信号交易{s.get('signal_trades',0)}笔)</div></div>
<div class="stat-card"><div class="stat-value">{s.get('profit_factor',0):.2f}</div><div class="stat-label">盈亏比</div></div>
<div class="stat-card"><div class="stat-value down">{s.get('max_drawdown_pct',0):.2f}%</div><div class="stat-label">最大回撤</div></div>
<div class="stat-card"><div class="stat-value">{s.get('sharpe_ratio',0):.2f}</div><div class="stat-label">夏普比率</div></div>
<div class="stat-card"><div class="stat-value">{s.get('avg_r_multiple',0):.2f}R</div><div class="stat-label">平均R倍数</div></div>
<div class="stat-card"><div class="stat-value">{s.get('filled_trades',0)}</div><div class="stat-label">成交/信号({s.get('total_signals',0)}) · 超时{s.get('time_exit_trades',0)}笔</div></div>
<div class="stat-card"><div class="stat-value">${s.get('final_capital',0):,.0f}</div><div class="stat-label">最终资金 (初始${s.get('initial_capital',0):,})</div></div>
<div class="stat-card"><div class="stat-value">{s.get('max_streak_win',0)}</div><div class="stat-label">最长连胜 · 最长连亏{s.get('max_streak_loss',0)}</div></div>
</div>

{signal_html}

<div class="chart-box">
<h3>资金权益曲线</h3>
<canvas id="equityChart" height="100"></canvas>
</div>

<div class="table-box">
<h3>按月统计</h3>
<table><thead><tr><th>月份</th><th>笔数</th><th>总盈亏</th><th>胜率</th></tr></thead>
<tbody>{monthly_rows}</tbody></table>
</div>

<div class="table-box">
<h3>交易明细（最近50笔）</h3>
<div style="overflow-x:auto">
<table><thead><tr>
<th>信号日</th><th>入场日</th><th>出场日</th><th>方向</th>
<th>入场价</th><th>出场价</th><th>止损</th><th>止盈</th><th>盈亏</th><th>R倍数</th><th>持仓</th><th>原因</th>
</tr></thead><tbody>{rows}</tbody></table>
</div></div>

<p class="footer">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
</div>

<script>
new Chart(document.getElementById('equityChart'),{{
type:'line',
data:{{labels:{eq_labels},datasets:[{{label:'权益',data:{eq_values},
borderColor:'#58a6ff',backgroundColor:'rgba(88,166,255,0.08)',
fill:true,tension:0.1,pointRadius:0}}]}},
options:{{responsive:true,
plugins:{{legend:{{labels:{{color:'#8b949e'}}}}}},
scales:{{x:{{ticks:{{color:'#8b949e',maxTicksLimit:20}}}},
y:{{ticks:{{color:'#8b949e',callback:v=>'$'+v.toFixed(0)}}}}}}}}
}});
</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════
#  入口
# ══════════════════════════════════════════════════════════════════

def main():
    cfg = Config("config.json")

    # 解析命令行参数
    symbol = sys.argv[1] if len(sys.argv) > 1 else cfg.SYMBOLS[0]
    start_day = int(sys.argv[2]) if len(sys.argv) > 2 else None
    end_day = int(sys.argv[3]) if len(sys.argv) > 3 else None

    print(f"\n{'='*60}")
    print(f"  AI 策略回测")
    print(f"{'='*60}")
    print(f"  交易对: {symbol}")
    print(f"  数据目录: {cfg.HISTORY_DIR}")
    print(f"  输出目录: {cfg.OUTPUT_DIR}")
    if start_day is not None:
        print(f"  自定义范围: day {start_day} → {end_day}")

    engine = BacktestEngine(cfg, symbol)
    t0 = time.time()
    report = engine.run()
    elapsed = time.time() - t0

    if "error" in report:
        print(f"\n❌ 错误: {report['error']}")
        return

    print_report(report)
    save_report(report, cfg)
    print(f"\n⏱ 总耗时: {elapsed:.0f} 秒 ({elapsed/60:.1f} 分钟)")


if __name__ == "__main__":
    main()
