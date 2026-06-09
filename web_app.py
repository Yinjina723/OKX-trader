# web_app.py
"""Web 面板：Flask 应用 —— 日线分析 + 回测"""

from __future__ import annotations

import json
import logging
import os
import threading
import uuid
from datetime import datetime, date
from typing import Any, Dict, Optional

from flask import Flask, jsonify, request, render_template, send_from_directory

from config import Config
from logger import setup_logger
from main import analyze_daily
from trade_log import write_signal, init_log_file

app = Flask(__name__, template_folder=os.path.join(os.path.dirname(__file__), "templates"))

# ── 回测任务状态（线程安全） ──
_backtest_lock = threading.Lock()
_backtest_tasks: Dict[str, Dict[str, Any]] = {}  # task_id → {status, progress, result, error, ...}

# ── 决策引擎状态 ──
_decision_lock = threading.Lock()
_decision_running = False


def load_config() -> Config:
    return Config("config.json")


def _ensure_logging():
    root = logging.getLogger()
    if not any(h for h in root.handlers if getattr(h, "baseFilename", None)):
        try:
            setup_logger(load_config())
        except Exception:
            pass


_ensure_logging()
logger = logging.getLogger(__name__)


def _try_write_trade_log(symbol: str, direction: str, ai: dict):
    """尝试将 AI 分析结果写入交易日志 Excel，失败不阻断主流程。"""
    try:
        # 确保日志文件存在
        init_log_file()

        if direction in ("long", "short"):
            entry = ai.get("entry") or 0
            sl = ai.get("stop_loss") or 0
            tp = ai.get("take_profit1") or 0
            note = ai.get("tomorrow_prediction", "")

            # 简单推算回测预期盈亏%（基于盈亏比）
            backtest_pct = None
            if entry > 0 and sl > 0 and tp > 0 and entry != sl:
                rr = abs(tp - entry) / abs(entry - sl)
                # 假设胜率50%，预期 = rr*50% - 1*50%
                backtest_pct = round((rr * 0.5 - 0.5) * 100 / entry * abs(entry - sl), 2)

            row = write_signal(
                symbol=symbol,
                direction=direction,
                entry=entry,
                stop_loss=sl,
                take_profit=tp,
                backtest_expect_pct=backtest_pct,
                note=note,
            )
            if row > 0:
                logger.info("交易日志已自动写入: %s %s 第%s行", symbol, direction, row)
            else:
                logger.warning("交易日志写入失败: %s", symbol)
        else:
            logger.info("方向为 neutral，不写入交易日志")
    except Exception as e:
        logger.warning("写入交易日志失败(非致命): %s", e)


@app.get("/")
def index() -> str:
    cfg = load_config()
    return render_template(
        "panel.html",
        symbols=cfg.SYMBOLS,
    )


@app.get("/backtest")
def backtest_page() -> str:
    """回测页面。"""
    cfg = load_config()
    return render_template(
        "backtest.html",
        symbols=cfg.SYMBOLS,
    )


@app.get("/api/config")
def api_get_config() -> Any:
    cfg_path = os.path.join(os.path.dirname(__file__), "config.json")
    with open(cfg_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    symbols = raw.get("SYMBOLS") or raw.get("SYMBOL", [])
    raw["SYMBOLS"] = symbols
    return jsonify(raw)


@app.post("/api/config")
def api_update_config() -> Any:
    payload: Dict[str, Any] = request.get_json(silent=True) or {}
    cfg_path = os.path.join(os.path.dirname(__file__), "config.json")
    with open(cfg_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    # 更新交易对
    symbols_str = payload.get("SYMBOLS")
    if isinstance(symbols_str, str):
        symbols = [s.strip() for s in symbols_str.split(",") if s.strip()]
        if symbols:
            raw["SYMBOL"] = symbols

    # 可更新的字段
    for key in ["DAILY_LOOKBACK", "WICK_SHADOW_RATIO", "ADVANCED_INDICATORS",
                "AI_TEMPERATURE", "AI_MAX_TOKENS", "PROXY_URL", "SITE"]:
        if key in payload and payload[key] is not None:
            raw[key] = payload[key]

    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(raw, f, ensure_ascii=False, indent=4)

    return jsonify({"message": "配置已保存"})


@app.post("/api/analyze")
def api_analyze() -> Any:
    """日线分析入口 —— 为指定 symbol 运行完整分析管道。"""
    data = request.get_json(silent=True) or {}
    symbol = data.get("symbol")
    if not symbol:
        return jsonify({"error": "symbol 必填"}), 400

    cfg = load_config()
    if symbol not in cfg.SYMBOLS:
        return jsonify({"error": f"symbol {symbol} 不在配置的 SYMBOLS 中"}), 400

    logger.info("开始日线分析: %s", symbol)
    try:
        result = analyze_daily(symbol=symbol, cfg=cfg)

        if result.get("error"):
            return jsonify(result), 500

        # 提取摘要
        y = result.get("yesterday", {})
        ai = result.get("ai_analysis", {})
        manip = result.get("manipulation", {})
        phase = manip.get("phase_result", {})

        direction = ai.get("direction", "neutral")
        strength = ai.get("strength", "medium")
        dir_cn = {"long": "做多", "short": "做空", "neutral": "观望"}.get(direction, "观望")
        str_cn = {"strong": "强", "medium": "中等", "weak": "较弱"}.get(strength, "中等")

        # 明日预测简述
        tomorrow_pred = ai.get("tomorrow_prediction", "")

        if direction == "neutral":
            if ai.get("ai_unavailable"):
                summary_cn = f"⚠️ DeepSeek 未配置，无法预测。阶段：{phase.get('phase_cn','?')}。"
            else:
                summary_cn = f"今日观望。{tomorrow_pred or '无明确方向'} 阶段：{phase.get('phase_cn','?')}。"
        else:
            entry = ai.get("entry") or 0
            sl = ai.get("stop_loss") or 0
            tp1 = ai.get("take_profit1") or 0
            tp2 = ai.get("take_profit2") or 0
            rr = abs(tp1 - entry) / abs(entry - sl) if entry and sl and tp1 and entry != sl else 0
            summary_cn = (
                f"【今日计划】{dir_cn} | 挂单{entry:.6f} | 止损{sl:.6f} | "
                f"止盈{tp1:.6f} | 盈亏比{rr:.1f}:1 | 强度{str_cn}"
            )
            if tomorrow_pred:
                summary_cn += f"\n预测: {tomorrow_pred}"

        # K线形态 + 背离 合并到一句话
        patterns = result.get("patterns", [])
        divergence = result.get("divergence", {})
        rsi_div = divergence.get("rsi", {}).get("detail", "")
        macd_div = divergence.get("macd", {}).get("detail", "")
        extra = []
        if patterns:
            extra.append("形态:" + ",".join(p["name"] for p in patterns[:3]))
        if rsi_div:
            extra.append(rsi_div.split(":")[0] if ":" in rsi_div else rsi_div)
        if macd_div:
            extra.append(macd_div.split(":")[0] if ":" in macd_div else macd_div)
        if extra:
            summary_cn += " | " + " ".join(extra)

        pattern_events = [{"type": p.get("direction","neutral"), "type_cn": p.get("name",""), "detail": p.get("name","")} for p in patterns]

        signal = {
            "direction": direction,
            "entry": ai.get("entry"),
            "stop_loss": ai.get("stop_loss"),
            "take_profit1": ai.get("take_profit1"),
            "take_profit2": ai.get("take_profit2"),
            "strength": strength,
            "market_state": ai.get("market_state", ""),
            "tomorrow_prediction": ai.get("tomorrow_prediction", ""),
            "key_support": ai.get("key_support"),
            "key_resistance": ai.get("key_resistance"),
            "ai_unavailable": ai.get("ai_unavailable", False),
            "manipulation": manip,
            "patterns": patterns,
            "divergence": divergence,
            "ma_alignment": result.get("ma_alignment", {}),
            "wicks": manip.get("wicks", []),
        }

        # ── 自动写入交易日志Excel ──
        _try_write_trade_log(symbol, direction, ai)

        return jsonify({
            "symbol": symbol,
            "yesterday_date": result.get("yesterday_date", ""),
            "current_price": result.get("current_price", 0),
            "yesterday": y,
            "summary_cn": summary_cn,
            "signal": signal,
            "kline": result.get("kline", []),
            "pattern_events": pattern_events,
            "indicators": result.get("indicators", {}),
            "liquidation_hunt": result.get("liquidation_hunt"),
            "capital_plan": result.get("capital_plan"),
        })

    except Exception as e:
        logger.exception("日线分析失败: %s", e)
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  回测 API
# ══════════════════════════════════════════════════════════════


@app.get("/api/backtest/data-status")
def api_backtest_data_status() -> Any:
    """检查历史数据是否已预下载。"""
    symbol = request.args.get("symbol", "")
    if not symbol:
        return jsonify({"error": "symbol 必填"}), 400

    cfg = load_config()
    safe = symbol.replace("-", "_")
    hist = cfg.HISTORY_DIR
    kline_path = os.path.join(hist, f"kline_{safe}.csv")
    funding_path = os.path.join(hist, f"funding_{safe}.csv")

    kline_exists = os.path.exists(kline_path)
    funding_exists = os.path.exists(funding_path)

    # 如果 K 线存在，读取行数
    kline_count = 0
    funding_count = 0
    date_range = ""
    if kline_exists:
        try:
            import pandas as pd
            df = pd.read_csv(kline_path, parse_dates=["timestamp"])
            # 去重后计数（旧数据可能有重复行）
            df = df.drop_duplicates(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
            kline_count = len(df)
            if kline_count > 0:
                date_range = f"{str(df['timestamp'].iloc[0])[:10]} → {str(df['timestamp'].iloc[-1])[:10]}"
        except Exception:
            pass
    if funding_exists:
        try:
            import pandas as pd
            fdf = pd.read_csv(funding_path, parse_dates=["fundingTime"])
            funding_count = len(fdf)
        except Exception:
            pass

    # ── 检查数据新鲜度 ──
    last_date = ""
    stale = False
    if kline_count > 0:
        try:
            last_ts = df["timestamp"].max()
            if hasattr(last_ts, 'strftime'):
                last_date = last_ts.strftime("%Y-%m-%d")
            else:
                last_date = str(last_ts)[:10]
            yesterday = date.today()
            if hasattr(last_ts, 'date'):
                stale = last_ts.date() < yesterday
            else:
                stale = pd.Timestamp(last_ts).date() < yesterday
        except Exception:
            pass

    return jsonify({
        "symbol": symbol,
        "kline_exists": kline_exists,
        "kline_count": kline_count,
        "funding_exists": funding_exists,
        "funding_count": funding_count,
        "date_range": date_range,
        "last_date": last_date,
        "stale": stale,
        "ready": kline_exists,  # 至少需要 K 线
    })


@app.post("/api/backtest/prepare")
def api_backtest_prepare() -> Any:
    """触发数据预下载（同步，很快）。"""
    data = request.get_json(silent=True) or {}
    symbol = data.get("symbol", "")

    if not symbol:
        return jsonify({"error": "symbol 必填"}), 400

    cfg = load_config()
    try:
        from okx_client import OKXClient
        from prepare_data import download_symbol

        client = OKXClient(cfg)
        ok = download_symbol(client, symbol, cfg.HISTORY_DIR)
        return jsonify({"success": ok, "symbol": symbol})
    except Exception as e:
        logger.exception("预下载失败: %s", e)
        return jsonify({"error": str(e)}), 500


@app.post("/api/backtest/start")
def api_backtest_start() -> Any:
    """启动回测任务（异步，返回 task_id 供轮询）。"""
    data = request.get_json(silent=True) or {}
    symbol = data.get("symbol", "")
    start_day = data.get("start_day")
    end_day = data.get("end_day")

    if not symbol:
        return jsonify({"error": "symbol 必填"}), 400

    cfg = load_config()
    task_id = str(uuid.uuid4())[:8]

    with _backtest_lock:
        _backtest_tasks[task_id] = {
            "status": "starting",
            "progress": 0,
            "progress_text": "初始化...",
            "symbol": symbol,
            "result": None,
            "error": None,
            "started_at": datetime.now().isoformat(),
        }

    # 在后台线程中运行回测
    thread = threading.Thread(
        target=_run_backtest_thread,
        args=(task_id, cfg, symbol, start_day, end_day),
        daemon=True,
    )
    thread.start()

    return jsonify({"task_id": task_id})


@app.get("/api/backtest/status")
def api_backtest_status() -> Any:
    """轮询回测任务状态。"""
    task_id = request.args.get("task_id", "")
    if not task_id:
        return jsonify({"error": "task_id 必填"}), 400

    with _backtest_lock:
        task = _backtest_tasks.get(task_id)

    if not task:
        return jsonify({"error": "任务不存在"}), 404

    resp = {
        "task_id": task_id,
        "status": task["status"],
        "progress": task["progress"],
        "progress_text": task.get("progress_text", ""),
        "symbol": task.get("symbol", ""),
        "error": task.get("error"),
    }

    if task["status"] == "done":
        resp["summary"] = task.get("result", {}).get("summary")
        resp["report_id"] = task.get("report_id")

    return jsonify(resp)


@app.get("/api/backtest/reports")
def api_backtest_reports() -> Any:
    """列出历史回测报告文件。"""
    cfg = load_config()
    out_dir = cfg.OUTPUT_DIR
    import glob
    reports = []
    for f in sorted(glob.glob(os.path.join(out_dir, "backtest_*.json")), reverse=True):
        fname = os.path.basename(f)
        try:
            with open(f, "r", encoding="utf-8") as fp:
                r = json.load(fp)
            reports.append({
                "filename": fname,
                "symbol": r.get("symbol", "?"),
                "period": r.get("period", {}),
                "summary": r.get("summary", {}),
            })
        except Exception:
            pass
    return jsonify(reports[:20])


@app.get("/api/backtest/report/<filename>")
def api_backtest_report_detail(filename: str) -> Any:
    """获取指定报告详情。"""
    cfg = load_config()
    filepath = os.path.join(cfg.OUTPUT_DIR, filename)
    if not os.path.exists(filepath):
        return jsonify({"error": "报告不存在"}), 404
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception:
        return jsonify({"error": "读取失败"}), 500


@app.get("/backtest-reports/<path:filename>")
def serve_backtest_html(filename: str) -> Any:
    """直接提供 OUTPUT_DIR 中的 HTML 报告文件。"""
    cfg = load_config()
    return send_from_directory(cfg.OUTPUT_DIR, filename)


def _run_backtest_thread(task_id: str, cfg: Config, symbol: str,
                         start_day: Optional[int], end_day: Optional[int]):
    """后台线程执行回测，更新 _backtest_tasks 状态。"""
    try:
        with _backtest_lock:
            _backtest_tasks[task_id]["status"] = "running"
            _backtest_tasks[task_id]["progress_text"] = "检查数据新鲜度..."

        # ── 自动刷新：缓存数据最后日期 < 昨天，先拉最新 ──
        safe = symbol.replace("-", "_")
        kline_path = os.path.join(cfg.HISTORY_DIR, f"kline_{safe}.csv")
        need_refresh = True

        if os.path.exists(kline_path):
            try:
                import pandas as pd
                df_existing = pd.read_csv(kline_path, parse_dates=["timestamp"])
                if len(df_existing) > 0:
                    last_ts = df_existing["timestamp"].max()
                    last_date = last_ts.date() if hasattr(last_ts, 'date') else pd.Timestamp(last_ts).date()
                    yesterday = date.today()
                    if last_date >= yesterday:
                        need_refresh = False
                        logger.info("数据已是最新 (%s)，跳过刷新", last_date.isoformat())
                    else:
                        logger.info("数据过期 (最后 %s)，自动拉取最新数据...", last_date.isoformat())
            except Exception:
                pass

        if need_refresh:
            with _backtest_lock:
                _backtest_tasks[task_id]["progress_text"] = "正在拉取最新数据..."
            from okx_client import OKXClient
            from prepare_data import download_symbol
            client = OKXClient(cfg)
            ok = download_symbol(client, symbol, cfg.HISTORY_DIR)
            if not ok:
                with _backtest_lock:
                    _backtest_tasks[task_id]["status"] = "error"
                    _backtest_tasks[task_id]["error"] = "数据下载失败，请检查网络或API密钥"
                return

        with _backtest_lock:
            _backtest_tasks[task_id]["progress_text"] = "加载数据..."

        from backtest import BacktestEngine

        engine = BacktestEngine(cfg, symbol)

        # 加载数据
        if not engine.load_data():
            with _backtest_lock:
                _backtest_tasks[task_id]["status"] = "error"
                _backtest_tasks[task_id]["error"] = "数据加载失败"
            return

        # 预计算指标
        with _backtest_lock:
            _backtest_tasks[task_id]["progress_text"] = "计算技术指标..."
        engine._compute_indicators()

        n = len(engine.df)
        actual_start = max(engine.warmup_days, start_day or engine.warmup_days)
        actual_end = min(n - 5, end_day or n - 5)

        if actual_start >= actual_end:
            with _backtest_lock:
                _backtest_tasks[task_id]["status"] = "error"
                _backtest_tasks[task_id]["error"] = (
                    f"数据不足：K线仅 {n} 根，至少需要 {engine.warmup_days + 6} 根"
                    f"（预热 {engine.warmup_days} 天 + 分析天数）"
                )
            return

        total_days = actual_end - actual_start + 1

        # 加载检查点
        engine._checkpoint_path = os.path.join(
            cfg.OUTPUT_DIR, f"bt_ckpt_{engine.safe_name}.json"
        )
        engine._load_checkpoint()

        pending = [i for i in range(actual_start, actual_end + 1)
                   if i not in engine.completed_days]
        completed_before = len(engine.completed_days)

        if not pending:
            # 所有天数已从检查点恢复，但仍需生成当前信号（双时间框架）
            _generate_current_signal_dual_tf(task_id, cfg, engine)
            report = engine._build_report()
            _save_backtest_result(task_id, cfg, engine, report)
            return

        import time
        from concurrent.futures import ThreadPoolExecutor, as_completed

        batch_size = engine.max_workers * 3
        total_pending = len(pending)

        for b_start in range(0, total_pending, batch_size):
            batch = pending[b_start:b_start + batch_size]

            with ThreadPoolExecutor(max_workers=engine.max_workers) as pool:
                futures = {pool.submit(engine._analyze_one_day, d): d
                          for d in batch}

                for fut in as_completed(futures):
                    day = futures[fut]
                    try:
                        analysis = fut.result(timeout=engine.ai_timeout)
                    except Exception:
                        engine.completed_days.add(day)
                        continue

                    if analysis:
                        trade = engine._simulate_trade(analysis)
                        if trade:
                            engine.trades.append(trade)

                    engine.completed_days.add(day)

                    # 更新进度
                    done = len(engine.completed_days) - completed_before
                    pct = min(99, int(done / total_pending * 100))
                    with _backtest_lock:
                        _backtest_tasks[task_id]["progress"] = pct
                        _backtest_tasks[task_id]["progress_text"] = (
                            f"分析中... {done}/{total_pending} 天 "
                            f"(已成交 {len([t for t in engine.trades if t.get('filled')])} 笔)"
                        )

            # 每批次保存检查点
            engine._save_checkpoint()

        # 生成报告
        with _backtest_lock:
            _backtest_tasks[task_id]["progress"] = 99
            _backtest_tasks[task_id]["progress_text"] = "生成当前信号..."

        # ── 生成当前点位预测（双时间框架：日线+今日1H）──
        with _backtest_lock:
            _backtest_tasks[task_id]["progress_text"] = "生成当前信号（双时间框架）..."
        _generate_current_signal_dual_tf(task_id, cfg, engine)

        with _backtest_lock:
            _backtest_tasks[task_id]["progress_text"] = "生成报告..."

        report = engine._build_report()
        _save_backtest_result(task_id, cfg, engine, report)

    except Exception as e:
        logger.exception("回测线程异常: %s", e)
        with _backtest_lock:
            _backtest_tasks[task_id]["status"] = "error"
            _backtest_tasks[task_id]["error"] = str(e)


def _generate_current_signal_dual_tf(task_id: str, cfg: Config, engine):
    """生成当前信号：双时间框架分析（日线 + 今日1H）。"""
    import pandas as pd
    from patterns import (
        detect_candlestick_patterns,
        detect_rsi_divergence,
        detect_macd_divergence,
        detect_ma_alignment,
    )
    from manipulation.daily_engine import run_daily_manipulation
    from ai_analysis import analyze_with_two_timeframes

    engine.current_signal = None
    engine._current_signal_error = None

    try:
        last_idx = len(engine.df) - 2
        min_required = max(15, engine.warmup_days)
        if last_idx < min_required:
            engine._current_signal_error = (
                f"数据不足: 需至少{min_required}天预热，当前仅{last_idx + 1}根K线"
            )
            return

        # Step 1: 计算日线上下文
        df_view = engine.df.iloc[:last_idx + 1].copy()
        patterns = detect_candlestick_patterns(df_view, lookback=10)
        rsi_div = detect_rsi_divergence(df_view)
        macd_div = detect_macd_divergence(df_view)
        divergence = {"rsi": rsi_div, "macd": macd_div}
        ma_alignment = detect_ma_alignment(df_view)
        manipulation = run_daily_manipulation(
            df_view, symbol=engine.symbol,
            wick_shadow_ratio=cfg.WICK_SHADOW_RATIO,
        )
        sentiment = engine._build_sentiment(last_idx)

        # Step 2: 拉取今日1H K线
        hourly_df = None
        from okx_client import OKXClient
        live_client = OKXClient(cfg)
        try:
            hourly_raw = live_client.get_klines(engine.symbol, bar="1H", limit=24)
            if hourly_raw:
                hourly_rows = OKXClient.parse_klines(hourly_raw)
                hourly_df = pd.DataFrame(hourly_rows)
                hourly_df["timestamp"] = pd.to_datetime(hourly_df["timestamp"], unit="ms")
                today_utc = pd.Timestamp.utcnow().date()
                hourly_df["date"] = hourly_df["timestamp"].dt.date
                hourly_today = hourly_df[hourly_df["date"] == today_utc].sort_values("timestamp")
                if len(hourly_today) > 0:
                    hourly_df = hourly_today
                else:
                    hourly_df = hourly_df.sort_values("timestamp").tail(24)
        except Exception as e:
            logger.warning("拉取1H K线失败: %s，将用纯日线分析", e)

        # Step 3: 双时间框架AI分析
        ai_result = analyze_with_two_timeframes(
            cfg, engine.symbol, engine.df, last_idx,
            hourly_df,
            manipulation=manipulation, patterns=patterns,
            divergence=divergence, ma_alignment=ma_alignment,
            sentiment=sentiment,
        )

        # Step 4: 获取实盘价格
        live_price = None
        try:
            ticker_result = live_client.get_ticker(engine.symbol)
            if ticker_result:
                live_price = float(ticker_result.get("last", 0))
        except Exception:
            pass

        engine.current_signal = {
            "day_idx": last_idx,
            "date": str(engine.df["timestamp"].iloc[last_idx])[:10],
            "close": float(engine.df["close"].iloc[last_idx]),
            "ai": ai_result,
            "live_price": live_price,
            "live_time": datetime.now().isoformat(),
            "_analysis_mode": "dual_timeframe",
        }

        logger.info("双时间框架信号: direction=%s entry=%s",
                    ai_result.get("direction"), ai_result.get("entry"))

    except Exception as e:
        engine._current_signal_error = f"生成当前信号异常: {e}"
        logger.warning("生成当前信号失败: %s", e)


def _save_backtest_result(task_id: str, cfg: Config, engine, report: Dict):
    """保存回测结果并更新任务状态。"""
    from backtest import save_report
    try:
        save_report(report, cfg)
    except Exception as e:
        logger.warning("报告保存失败: %s", e)

    # 找到最新生成的文件
    import glob
    html_files = sorted(
        glob.glob(os.path.join(cfg.OUTPUT_DIR, f"backtest_{engine.safe_name}_*.html")),
        reverse=True,
    )
    json_files = sorted(
        glob.glob(os.path.join(cfg.OUTPUT_DIR, f"backtest_{engine.safe_name}_*.json")),
        reverse=True,
    )

    with _backtest_lock:
        _backtest_tasks[task_id]["status"] = "done"
        _backtest_tasks[task_id]["progress"] = 100
        _backtest_tasks[task_id]["progress_text"] = "完成"
        _backtest_tasks[task_id]["result"] = report
        _backtest_tasks[task_id]["report_id"] = (
            os.path.basename(html_files[0]) if html_files else None
        )
        _backtest_tasks[task_id]["report_json"] = (
            os.path.basename(json_files[0]) if json_files else None
        )


# ══════════════════════════════════════════════════════════════
#  导出交易跟踪表
# ══════════════════════════════════════════════════════════════

@app.get("/api/export/trade-journal")
def api_export_trade_journal():
    """生成并下载实盘交易跟踪 Excel 表。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HEADER_FILL2 = PatternFill("solid", fgColor="16213e")
    BLUE_FONT = Font(name="Arial", color="0066CC", size=10)
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def hdr(ws, row, n, fill=None):
        for c in range(1, n + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEADER_FONT
            cell.fill = fill or HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

    def drow(ws, row, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = CENTER
            cell.border = BORDER

    cfg = load_config()
    symbols = cfg.SYMBOLS or ["ALLO-USDT-SWAP"]

    # ── Sheet 1: 交易规则 ──
    ws1 = wb.active
    ws1.title = "交易规则"
    ws1.sheet_properties.tabColor = "16213e"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "AI策略实盘跟踪表"
    ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="1a1a2e")
    ws1["A1"].alignment = CENTER
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A2:F2")
    sym_str = " / ".join(symbols[:3])
    ws1["A2"] = f"本金 $220  |  杠杆 1x  |  分仓 50/50  |  每笔保证金 $10~15  |  币种: {sym_str}"
    ws1["A2"].font = Font(name="Arial", size=9, color="888888")
    ws1["A2"].alignment = CENTER

    rules = [
        ("账户参数", ""),
        ("初始入金", "$220"),
        ("杠杆倍数", "1x（资金卫士模式）"),
        ("分仓比例", "50% 首段入场 + 50% 备用金"),
        ("每笔保证金", "$10~$15（本金 5%~7%）"),
        ("名义仓位", "$10~$15（保证金×1x杠杆）"),
        ("单笔最大亏损", "≈$5（止损触发时）"),
        ("最多同时持仓", "1 笔"),
        ("", ""),
        ("纪律铁规", ""),
        ("❶ 每天最多1笔", "与回测逻辑一致"),
        ("❷ 止损绝不手动撤", "触发就认，这是验证系统"),
        ("❸ 超时15天平仓", "对齐回测 max_trade_days=15"),
        ("❹ 不加仓不补仓", "回测没这逻辑，补了无法对比"),
        ("❺ 连亏5笔就停", "实盘与回测差距大，先排查"),
        ("❻ 每笔必填表", "方便和回测报告交叉对比"),
    ]
    row = 4
    for label, val in rules:
        if label in ("账户参数", "纪律铁规"):
            ws1.merge_cells(f"A{row}:F{row}")
            ws1.cell(row=row, column=1, value=f"▎{label}")
            ws1.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=11, color="e76f51")
            ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FFF3E0")
            row += 1
            continue
        if not label:
            row += 1
            continue
        ws1.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws1.cell(row=row, column=1).alignment = LEFT
        ws1.merge_cells(f"B{row}:F{row}")
        ws1.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10, color="333333")
        ws1.cell(row=row, column=2).alignment = LEFT
        for c in range(1, 7):
            ws1.cell(row=row, column=c).border = BORDER
        row += 1
    ws1.column_dimensions["A"].width = 20
    for c in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[c].width = 16

    # ── Sheet 2: 交易日志 ──
    ws2 = wb.create_sheet("交易日志")
    ws2.sheet_properties.tabColor = "2a9d8f"
    headers = [
        "序号", "信号日", "币种", "方向", "AI入场价", "AI止损", "AI止盈",
        "保证金($)", "杠杆", "名义仓位($)", "实盘入场日", "实盘入场价",
        "实盘出场日", "实盘出场价", "实盘盈亏($)", "实盘盈亏(%)",
        "出场原因", "持仓天数", "回测预期盈亏(%)", "偏差(%)", "备注"
    ]
    N = len(headers)
    widths = [5, 11, 20, 6, 12, 12, 12, 10, 6, 12, 11, 12, 11, 12, 12, 12, 12, 8, 14, 10, 20]
    ws2.merge_cells(f"A1:{get_column_letter(N)}1")
    ws2["A1"] = "交易日志 — 每笔必须完整填写"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="1a1a2e")
    ws2["A1"].alignment = CENTER
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws2.cell(row=2, column=i, value=h)
        ws2.column_dimensions[get_column_letter(i)].width = w
    hdr(ws2, 2, N)
    ws2.freeze_panes = "A3"

    for row in range(3, 53):
        ws2.cell(row=row, column=1, value=row - 2)
        ws2.cell(row=row, column=10, value=f'=IF(H{row}="","",H{row}*I{row})')
        ws2.cell(row=row, column=15, value=(
            f'=IF(OR(L{row}="",N{row}="",D{row}=""),"",'
            f'IF(D{row}="long",(N{row}-L{row})/L{row}*J{row},'
            f'(L{row}-N{row})/L{row}*J{row}))'
        ))
        ws2.cell(row=row, column=16, value=f'=IF(H{row}="","",O{row}/H{row}*100)')
        ws2.cell(row=row, column=18, value=f'=IF(OR(K{row}="",M{row}=""),"",DAYS(M{row},K{row}))')
        ws2.cell(row=row, column=20, value=f'=IF(OR(P{row}="",S{row}=""),"",P{row}-S{row})')
        drow(ws2, row, N)

    for row in range(3, 53):
        for c in [5, 6, 7, 12, 14]:
            ws2.cell(row=row, column=c).number_format = '0.000000'
        ws2.cell(row=row, column=8).number_format = '0.00'
        ws2.cell(row=row, column=10).number_format = '0.00'
        ws2.cell(row=row, column=15).number_format = '+0.00;-0.00'
        ws2.cell(row=row, column=16).number_format = '0.00'
        ws2.cell(row=row, column=19).number_format = '0.00'
        ws2.cell(row=row, column=20).number_format = '+0.00;-0.00'

    # Summary
    R = 54
    ws2.merge_cells(f"A{R}:G{R}")
    ws2[f"A{R}"] = "自动汇总统计"
    ws2[f"A{R}"].font = Font(name="Arial", bold=True, size=11, color="1a1a2e")
    ws2[f"A{R}"].fill = PatternFill("solid", fgColor="D6EAF8")

    summary = [
        ("总交易笔数", '=COUNTA(B3:B52)', '0'),
        ("成交笔数", '=COUNTIF(Q3:Q52,"<>未成交")-COUNTIF(Q3:Q52,"")', '0'),
        ("盈利笔数", '=COUNTIF(P3:P52,">0")', '0'),
        ("胜率", '=IF(B55-B57=0,"",D55/(B55-B57))', '0.0%'),
        ("总盈亏($)", '=SUM(O3:O52)', '+$#,##0.00;-$#,##0.00'),
        ("平均盈亏(%)", '=IF(B55=0,"",AVERAGE(P3:P52))', '0.00'),
        ("盈亏比", '=IF(AND(D55>0,E55>0),SUMIF(P3:P52,">0")/ABS(SUMIF(P3:P52,"<0")),"")', '0.00'),
        ("当前资金($)", '=220+SUM(O3:O52)', '$#,##0.00'),
        ("总收益率(%)", '=SUM(O3:O52)/220*100', '0.0%'),
    ]
    for i, (label, formula, fmt) in enumerate(summary):
        r = R + 1 + i
        ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=r, column=1).border = BORDER
        ws2.merge_cells(f"B{r}:G{r}")
        ws2.cell(row=r, column=2, value=formula).font = BLUE_FONT
        ws2.cell(row=r, column=2).border = BORDER
        ws2.cell(row=r, column=2).alignment = LEFT
        ws2.cell(row=r, column=2).number_format = fmt

    # ── Sheet 3: 统计分析 ──
    ws3 = wb.create_sheet("统计分析")
    ws3.sheet_properties.tabColor = "e76f51"
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "实盘 vs 回测 对比"
    ws3["A1"].font = Font(name="Arial", bold=True, size=13, color="1a1a2e")
    ws3["A1"].alignment = CENTER

    ws3["A3"] = "▎整体对比（回测值需手动填入）"
    ws3["A3"].font = Font(name="Arial", bold=True, size=11, color="e76f51")
    comp_h = ["指标", "回测值", "实盘值", "偏差", "判断"]
    for i, h in enumerate(comp_h, 1):
        ws3.cell(row=4, column=i, value=h)
    hdr(ws3, 4, 5, HEADER_FILL2)

    comp = [
        ("胜率(%)", "", "='交易日志'!B60", "=C5-B5", '=IF(ABS(D5)<=10,"✅ 达标","⚠ 偏差大")'),
        ("平均盈亏(%)", "", "='交易日志'!B62", "=C6-B6", '=IF(ABS(D6)<=5,"✅ 达标","⚠ 偏差大")'),
        ("盈亏比", "", "='交易日志'!B65", "=C7-B7", '=IF(C7>=1,"✅","❌ <1")'),
        ("总收益率(%)", "", "='交易日志'!B67", "=C8-B8", '=IF(ABS(D8)<=10,"✅ 达标","⚠ 偏差大")'),
    ]
    for i, (label, bt, real, dev, judge) in enumerate(comp):
        r = 5 + i
        ws3.cell(row=r, column=1, value=label)
        ws3.cell(row=r, column=2, value=bt if bt else "（手动填）")
        ws3.cell(row=r, column=3, value=real)
        ws3.cell(row=r, column=4, value=dev)
        ws3.cell(row=r, column=5, value=judge)
        drow(ws3, r, 5)
        ws3.cell(row=r, column=2).font = BLUE_FONT

    # By symbol
    R2 = 11
    ws3[f"A{R2}"] = "▎按币种统计"
    ws3[f"A{R2}"].font = Font(name="Arial", bold=True, size=11, color="e76f51")
    sym_h = ["币种", "笔数", "盈利", "胜率", "总盈亏(%)", "均盈亏(%)", "回测胜率(%)", "偏差(%)"]
    for i, h in enumerate(sym_h, 1):
        ws3.cell(row=R2 + 1, column=i, value=h)
    hdr(ws3, R2 + 1, 8, HEADER_FILL2)

    for i, coin in enumerate(symbols[:3]):
        r = R2 + 2 + i
        ws3.cell(row=r, column=1, value=coin).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=r, column=2, value=f'=COUNTIF(交易日志!C3:C52,"{coin}")')
        ws3.cell(row=r, column=3, value=f'=COUNTIFS(交易日志!C3:C52,"{coin}",交易日志!P3:P52,">0")')
        ws3.cell(row=r, column=4, value=f'=IF(B{r}=0,"",C{r}/B{r})')
        ws3.cell(row=r, column=5, value=f'=SUMIF(交易日志!C3:C52,"{coin}",交易日志!P3:P52)')
        ws3.cell(row=r, column=6, value=f'=IF(B{r}=0,"",AVERAGEIF(交易日志!C3:C52,"{coin}",交易日志!P3:P52))')
        ws3.cell(row=r, column=7, value="手动填")
        ws3.cell(row=r, column=8, value=f'=IF(OR(G{r}="手动填",G{r}=""),"",F{r}-G{r})')
        for c in range(1, 9):
            ws3.cell(row=r, column=c).border = BORDER
            ws3.cell(row=r, column=c).font = Font(name="Arial", size=10)
            ws3.cell(row=r, column=c).alignment = CENTER
        ws3.cell(row=r, column=4).number_format = '0.0%'
        ws3.cell(row=r, column=6).number_format = '0.00'
        ws3.cell(row=r, column=8).number_format = '0.00'

    # By direction
    R3 = R2 + 7
    ws3[f"A{R3}"] = "▎按方向统计"
    ws3[f"A{R3}"].font = Font(name="Arial", bold=True, size=11, color="e76f51")
    for i, h in enumerate(sym_h, 1):
        ws3.cell(row=R3 + 1, column=i, value=h)
    hdr(ws3, R3 + 1, 8, HEADER_FILL2)

    for i, d in enumerate(["long", "short"]):
        r = R3 + 2 + i
        ws3.cell(row=r, column=1, value=d).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=r, column=2, value=f'=COUNTIF(交易日志!D3:D52,"{d}")')
        ws3.cell(row=r, column=3, value=f'=COUNTIFS(交易日志!D3:D52,"{d}",交易日志!P3:P52,">0")')
        ws3.cell(row=r, column=4, value=f'=IF(B{r}=0,"",C{r}/B{r})')
        ws3.cell(row=r, column=5, value=f'=SUMIF(交易日志!D3:D52,"{d}",交易日志!P3:P52)')
        ws3.cell(row=r, column=6, value=f'=IF(B{r}=0,"",AVERAGEIF(交易日志!D3:D52,"{d}",交易日志!P3:P52))')
        ws3.cell(row=r, column=7, value="手动填")
        ws3.cell(row=r, column=8, value=f'=IF(OR(G{r}="手动填",G{r}=""),"",F{r}-G{r})')
        for c in range(1, 9):
            ws3.cell(row=r, column=c).border = BORDER
            ws3.cell(row=r, column=c).font = Font(name="Arial", size=10)
            ws3.cell(row=r, column=c).alignment = CENTER
        ws3.cell(row=r, column=4).number_format = '0.0%'
        ws3.cell(row=r, column=6).number_format = '0.00'
        ws3.cell(row=r, column=8).number_format = '0.00'

    # ── Sheet 4: 资金曲线 ──
    ws4 = wb.create_sheet("资金曲线")
    ws4.sheet_properties.tabColor = "264653"
    ws4.merge_cells("A1:D1")
    ws4["A1"] = "资金曲线"
    ws4["A1"].font = Font(name="Arial", bold=True, size=13, color="1a1a2e")
    ws4["A1"].alignment = CENTER

    for i, h in enumerate(["日期", "本笔盈亏($)", "累计盈亏($)", "账户余额($)"], 1):
        ws4.cell(row=2, column=i, value=h)
    hdr(ws4, 2, 4, HEADER_FILL2)

    ws4.cell(row=3, column=1, value="起始")
    ws4.cell(row=3, column=2, value=0)
    ws4.cell(row=3, column=3, value=0)
    ws4.cell(row=3, column=4, value=220)
    for c in range(1, 5):
        ws4.cell(row=3, column=c).font = Font(name="Arial", size=10)
        ws4.cell(row=3, column=c).alignment = CENTER
        ws4.cell(row=3, column=c).border = BORDER

    for row in range(4, 53):
        ws4.cell(row=row, column=1, value="（逐笔填）")
        ws4.cell(row=row, column=3, value=f'=IF(B{row}="","",C{row-1}+B{row})')
        ws4.cell(row=row, column=4, value=f'=220+C{row}')
        for c in range(1, 5):
            ws4.cell(row=row, column=c).font = Font(name="Arial", size=10)
            ws4.cell(row=row, column=c).alignment = CENTER
            ws4.cell(row=row, column=c).border = BORDER
        ws4.cell(row=row, column=2).number_format = '+0.00;-0.00'
        ws4.cell(row=row, column=3).number_format = '+0.00;-0.00'
        ws4.cell(row=row, column=4).number_format = '0.00'

    ws4.column_dimensions["A"].width = 14
    for c in ["B", "C", "D"]:
        ws4.column_dimensions[c].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="AI策略实盘跟踪表.xlsx")


# ══════════════════════════════════════════════════════════════
#  智能决策 API
# ══════════════════════════════════════════════════════════════


@app.get("/api/decision/latest")
def api_decision_latest() -> Any:
    """获取最新决策推荐结果（从缓存读取）。"""
    cfg = load_config()
    from decision_engine import load_decision_cache, run_decision_pipeline, save_decision_cache

    cache = load_decision_cache(cfg)
    if cache is None:
        return jsonify({"error": "暂无决策数据，请点击刷新", "results": []})

    return jsonify(cache)


@app.post("/api/decision/refresh")
def api_decision_refresh() -> Any:
    """手动触发决策刷新（同步或异步）。"""
    global _decision_running

    with _decision_lock:
        if _decision_running:
            return jsonify({"success": False, "error": "决策刷新已在进行中，请稍候..."})
        _decision_running = True

    try:
        logger.info("手动触发决策刷新...")
        cfg = load_config()

        from decision_engine import run_decision_pipeline, save_decision_cache

        # 运行管道（扫描配置中的 SYMBOLS）
        results = run_decision_pipeline(cfg)
        save_decision_cache(results, cfg)

        return jsonify({"success": True, "message": "决策刷新完成", "count": len(results)})
    except Exception as e:
        logger.exception("决策刷新异常: %s", e)
        return jsonify({"success": False, "error": str(e)})
    finally:
        with _decision_lock:
            _decision_running = False


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8488)
